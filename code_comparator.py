import sys
import os
import pandas as pd
import io
import csv  # ← NUOVO: necessario per gestire le virgolette
import json
import shutil
import logging
from pathlib import Path


from PyQt6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, 
                             QHBoxLayout, QTableWidget, QTableWidgetItem, 
                             QPushButton, QLabel, QHeaderView, QTabWidget,
                             QMessageBox, QFileDialog, QMenu, QDialog, 
                             QLineEdit, QComboBox, QFormLayout, QDialogButtonBox,
                             QGroupBox, QScrollArea, QCheckBox, 
                             QListWidget, QListWidgetItem, QRadioButton, QButtonGroup)
                             # ✅ AGGIUNTI: QListWidget, QListWidgetItem, QRadioButton, QButtonGroup
from PyQt6.QtCore import Qt, QStandardPaths
from PyQt6.QtGui import QColor, QKeySequence, QFont
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter

# Log configuration (errors only, visible in console)
logging.basicConfig(
    level=logging.ERROR,
    format='%(asctime)s - %(levelname)s - %(message)s'
)


# ============================================================
# GLOBAL EXCEPTION HANDLER
# ============================================================
def global_exception_handler(exctype, value, tb):
    """Gestore di eccezioni globale per errori non gestiti"""
    logging.exception("Errore non gestito:", exc_info=(exctype, value, tb))
    
    # Mostra dialog solo se QApplication è attiva
    if QApplication.instance():
        QMessageBox.critical(
            None, 
            "⌠Errore Critico",
            f"Si è verificato un errore imprevisto:\n\n{str(value)}\n\n"
            "L'applicazione potrebbe non funzionare correttamente."
        )
    
    # Chiama il gestore predefinito per mantenere il traceback
    sys.__excepthook__(exctype, value, tb)

sys.excepthook = global_exception_handler


# ============================================================
# GLOBAL CONSTANTS
# ============================================================
# Colors
HEADER_BG_COLOR = "#ecf0f1"
HEADER_TEXT_COLOR = "#2c3e50"
HEADER_BORDER_COLOR = "#bdc3c7"
HIGHLIGHT_COLOR = "#fff3cd"
MATCH_COLOR = "#e8f8f5"
MISMATCH_COLOR = "#fdedec"
WHITE_COLOR = "#ffffff"
DUPLICATE_HIGHLIGHT = "#ffeaa7"
DUPLICATE_KEY_COLOR = "#ff7675"

# Dimensions
BUTTON_HEIGHT = 35
COMPARE_BUTTON_HEIGHT = 60
SETTINGS_BUTTON_WIDTH = 140
SETTINGS_BUTTON_HEIGHT = 40
DEFAULT_ROW_HEIGHT = 30
HEADER_ROW_HEIGHT = 19

# Excel Export Column Widths (keyword-based mapping)
EXCEL_COLUMN_WIDTHS = {
    65: ["TITOLO", "TITLE", "DESCRIZIONE", "DESCRIPTION", "ISBD"],
    25: ["NOTE", "NOTES", "LEGAMI", "LINKS", "COMMENTI", "COMMENTS"],
    23: ["AUTORE", "AUTHOR", "EDITORE", "PUBLISHER", "COLLOCAZIONE", "LOCATION"],
    15: ["ISBN", "CODICE", "CODE", "INVENTARIO", "INVENTORY", "ID"],
    12: ["ANNO", "YEAR", "PREZZO", "PRICE", "SEQUENZA", "SEZIONE", "SECTION", "SPECIFICAZIONE"]
}

MAX_COLUMNS = 7
DEFAULT_EXCEL_MIN_ROWS = 100  # Righe minime in Excel export
DEFAULT_ZOOM_LEVEL = 110  # Zoom Excel
MIN_COLUMNS_PARSE_THRESHOLD = 0.5  # Soglia parsing (50% colonne minime)

# ============================================================
# GLOBAL CSS STYLES
# ============================================================
APP_STYLESHEET = """
    QMainWindow {
        background-color: #f5f6fa;
    }
    
    /* STILE BASE PULSANTI */
    QPushButton {
        background-color: #3498db;
        color: white;
        border: none;
        padding: 10px 20px;
        font-weight: bold;
        border-radius: 5px;
        font-size: 13px;
    }
    QPushButton:hover {
        background-color: #2980b9;
    }
    QPushButton:pressed {
        background-color: #21618c;
    }
    QPushButton:disabled {
        background-color: #bdc3c7;
        color: #7f8c8d;
    }
    
    /* PULSANTI SPECIALIZZATI (solo colori diversi) */
    QPushButton#btn_compare {
        background-color: #9b59b6;
        color: white;
        font-size: 16px;
        border: 3px solid #6c3483;
    }
    QPushButton#btn_compare:hover {
        background-color: #8e44ad;
        border: 3px solid #5b2c6f;
    }

    QPushButton#btn_export_matches {
        background-color: #27ae60;
        color: white;
    }
    QPushButton#btn_export_matches:hover {
        background-color: #229954;
    }

    QPushButton#btn_export_mismatches {
        background-color: #e67e22;
        color: white;
    }
    QPushButton#btn_export_mismatches:hover {
        background-color: #d35400;
    }

    QPushButton#btn_clear_results {
        background-color: #95a5a6;
        color: white;
    }
    QPushButton#btn_clear_results:hover {
        background-color: #7f8c8d;
    }

    QPushButton#btn_settings {
        background-color: #34495e;
        color: white;
    }
    QPushButton#btn_settings:hover {
        background-color: #2c3e50;
    }

    QPushButton#btn_swap {
        background-color: #f39c12;
        color: white;
    }
    QPushButton#btn_swap:hover {
        background-color: #e67e22;
    }
    
    /* TABELLE */
    QTableWidget {
        border: 1px solid #bdc3c7;
        border-radius: 5px;
        background-color: white;
        gridline-color: #ecf0f1;
        selection-background-color: #cce5ff;
        selection-color: #000000;
    }
    
    QHeaderView::section {
        background-color: #ecf0f1;
        color: #2c3e50;
        padding: 10px;
        border: 1px solid #bdc3c7;
        font-weight: bold;
        font-size: 13px;
    }
    QHeaderView::section:hover {
        background-color: #d5dbdb;
    }
    
    /* TABS */
    QTabWidget::pane {
        border: 1px solid #bdc3c7;
        border-radius: 5px;
        background-color: white;
    }
    QTabBar::tab {
        background-color: #ecf0f1;
        color: #2c3e50;
        padding: 10px 20px;
        border-top-left-radius: 5px;
        border-top-right-radius: 5px;
        margin-right: 2px;
        font-weight: bold;
    }
    QTabBar::tab:selected {
        background-color: #3498db;
        color: white;
    }
    QTabBar::tab:hover {
        background-color: #bdc3c7;
    }
"""


class SettingsDialog(QDialog):
    """
    Interfaccia Master-Detail con Radio Button per colonna chiave.
    Versione DEFINITIVA con bugfix critici.
    """
    def __init__(self, presets, current_preset, default_preset, protected_presets, parent=None):
        super().__init__(parent)
        self.presets = presets.copy()
        self.current_selected_name = current_preset
        self.default_preset = default_preset
        self.protected_presets = protected_presets
        self.result = None
        self.is_new_preset_mode = False
        
        self.setWindowTitle("⚙️ Gestione Preset di Configurazione")
        self.resize(950, 650)
        self.setMinimumSize(900, 600)
        
        self.init_ui()
        self.load_preset_list()
        self.select_preset_in_list(current_preset)

    @staticmethod
    def validate_preset_data(preset_name, key_column, columns):
        """Validazione unificata"""
        if not preset_name: 
            return False, "Il nome del preset è obbligatorio"
        if not key_column: 
            return False, "Seleziona una colonna come chiave (clicca sul radio button)"
        if len(columns) < 1: 
            return False, "Definisci almeno una colonna"
        if len(columns) > MAX_COLUMNS: 
            return False, f"Massimo {MAX_COLUMNS} colonne consentite"
        if key_column not in columns: 
            return False, f"La colonna chiave '{key_column}' deve essere presente nell'elenco"
        return True, ""

    def init_ui(self):
        main_layout = QHBoxLayout(self)
        main_layout.setSpacing(0)
        main_layout.setContentsMargins(0, 0, 0, 0)

        # ==========================================
        # LATO SINISTRO: LISTA
        # ==========================================
        left_widget = QWidget()
        left_widget.setStyleSheet("""
            QWidget {
                background-color: #f8f9fa; 
                border-right: 2px solid #dee2e6;
            }
        """)
        left_widget.setMinimumWidth(280)
        left_layout = QVBoxLayout(left_widget)
        left_layout.setContentsMargins(15, 15, 15, 15)
        left_layout.setSpacing(10)
        
        title_list = QLabel("📋 I TUOI PRESET")
        title_list.setStyleSheet("""
            font-weight: bold; 
            font-size: 13px;
            color: #2c3e50; 
            padding: 8px;
            background-color: #e9ecef;
            border-radius: 4px;
        """)
        left_layout.addWidget(title_list)

        self.label_preset_count = QLabel()
        self.label_preset_count.setStyleSheet("color: #6c757d; font-size: 11px; padding: 2px 8px;")
        left_layout.addWidget(self.label_preset_count)

        from PyQt6.QtWidgets import QListWidget
        self.list_presets = QListWidget()
        self.list_presets.setStyleSheet("""
            QListWidget {
                border: 2px solid #ced4da;
                border-radius: 5px;
                background-color: white;
                font-size: 12px;
            }
            QListWidget::item {
                padding: 14px 10px;
                border-bottom: 1px solid #f1f3f5;
            }
            QListWidget::item:hover {
                background-color: #f8f9fa;
            }
            QListWidget::item:selected {
                background-color: #e7f1ff;
                color: #084298;
                font-weight: bold;
                border-left: 4px solid #0d6efd;
            }
        """)
        self.list_presets.currentItemChanged.connect(self.on_selection_changed)
        left_layout.addWidget(self.list_presets)

        self.label_selection_info = QLabel()
        self.label_selection_info.setStyleSheet("""
            color: #6c757d; 
            font-size: 10px; 
            padding: 5px 8px;
            background-color: #f8f9fa;
            border-radius: 3px;
        """)
        self.label_selection_info.setWordWrap(True)
        left_layout.addWidget(self.label_selection_info)

        # Pulsanti di gestione (su 2 righe per spazio)
        left_buttons_top = QHBoxLayout()
        left_buttons_top.setSpacing(8)
        
        self.btn_add_new = QPushButton("➕ Nuovo")
        self.btn_add_new.setStyleSheet("""
            QPushButton {
                background-color: #27ae60;
                color: white;
                padding: 10px;
                font-weight: bold;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #229954;
            }
        """)
        self.btn_add_new.clicked.connect(self.prepare_new_preset)
        
        self.btn_delete = QPushButton("🗑️ Elimina")
        self.btn_delete.setStyleSheet("""
            QPushButton {
                background-color: #e74c3c;
                color: white;
                padding: 10px;
                font-weight: bold;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #c0392b;
            }
            QPushButton:disabled {
                background-color: #95a5a6;
            }
        """)
        self.btn_delete.clicked.connect(self.delete_selected_preset)
        
        left_buttons_top.addWidget(self.btn_add_new)
        left_buttons_top.addWidget(self.btn_delete)
        left_layout.addLayout(left_buttons_top)
        
        # Seconda riga: pulsanti di ordinamento
        left_buttons_bottom = QHBoxLayout()
        left_buttons_bottom.setSpacing(8)
        
        self.btn_move_up = QPushButton("⬆️ Sposta Su")
        self.btn_move_up.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                color: white;
                padding: 10px;
                font-weight: bold;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
            QPushButton:disabled {
                background-color: #95a5a6;
            }
        """)
        self.btn_move_up.clicked.connect(self.move_preset_up)
        self.btn_move_up.setEnabled(False)
        
        self.btn_move_down = QPushButton("⬇️ Sposta Giù")
        self.btn_move_down.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                color: white;
                padding: 10px;
                font-weight: bold;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
            QPushButton:disabled {
                background-color: #95a5a6;
            }
        """)
        self.btn_move_down.clicked.connect(self.move_preset_down)
        self.btn_move_down.setEnabled(False)
        
        left_buttons_bottom.addWidget(self.btn_move_up)
        left_buttons_bottom.addWidget(self.btn_move_down)
        left_layout.addLayout(left_buttons_bottom)
        
        main_layout.addWidget(left_widget)

        # ==========================================
        # LATO DESTRO: EDITOR
        # ==========================================
        right_widget = QWidget()
        right_widget.setStyleSheet("background-color: white;")
        right_layout = QVBoxLayout(right_widget)
        right_layout.setContentsMargins(25, 20, 25, 20)
        right_layout.setSpacing(15)

        self.editor_header = QLabel("✏️ Modifica Configurazione")
        self.editor_header.setStyleSheet("""
            font-size: 16px;
            font-weight: bold;
            color: #2c3e50;
            padding: 10px;
            border-bottom: 2px solid #3498db;
        """)
        right_layout.addWidget(self.editor_header)

        self.label_error = QLabel()
        self.label_error.setStyleSheet("""
            color: #dc3545; 
            background-color: #f8d7da;
            border: 1px solid #f5c2c7;
            border-radius: 4px;
            padding: 8px;
            font-size: 11px;
        """)
        self.label_error.setWordWrap(True)
        self.label_error.hide()
        right_layout.addWidget(self.label_error)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("QScrollArea { border: none; background-color: white; }")
        
        form_container = QWidget()
        form_layout = QFormLayout(form_container)
        form_layout.setLabelAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        form_layout.setSpacing(15)
        form_layout.setContentsMargins(10, 10, 10, 10)

        # Nome Preset
        name_label = QLabel("Nome Preset:")
        name_label.setStyleSheet("font-weight: bold; color: #495057;")
        
        self.input_name = QLineEdit()
        self.input_name.setPlaceholderText("Nome identificativo del preset")
        self.input_name.setStyleSheet("""
            QLineEdit {
                padding: 10px;
                border: 2px solid #ced4da;
                border-radius: 5px;
                font-size: 13px;
            }
            QLineEdit:focus {
                border: 2px solid #3498db;
            }
            QLineEdit:disabled {
                background-color: #e9ecef;
            }
        """)
        self.input_name.textChanged.connect(self.clear_error)
        self.input_name.setToolTip("Nome che identifica questo preset (es: 'Acquisti 2025', 'Stampa Registri')")
        
        form_layout.addRow(name_label, self.input_name)

        # Separatore
        separator = QLabel()
        separator.setStyleSheet("""
            background-color: #dee2e6; 
            min-height: 2px; 
            max-height: 2px;
            margin: 10px 0px;
        """)
        form_layout.addRow(separator)

        # Intestazione colonne
        columns_header_widget = QWidget()
        columns_header_layout = QVBoxLayout(columns_header_widget)
        columns_header_layout.setSpacing(5)
        columns_header_layout.setContentsMargins(0, 0, 0, 0)
        
        columns_title = QLabel("📊 Definizione Colonne")
        columns_title.setStyleSheet("""
            font-weight: bold; 
            font-size: 13px;
            color: #2c3e50;
        """)
        
        columns_instructions = QLabel(
            "🔑 <b>Seleziona la colonna chiave</b> cliccando sul radio button.<br>"
            "Il radio button si attiva automaticamente quando compili il campo.<br>"
            "La colonna chiave è quella usata per confrontare le liste (es: ISBN, Codice Inventario)"
        )
        columns_instructions.setStyleSheet("""
            color: #6c757d; 
            font-size: 11px;
            background-color: #fff3cd;
            border: 1px solid #ffc107;
            border-radius: 4px;
            padding: 8px;
        """)
        columns_instructions.setWordWrap(True)
        
        columns_header_layout.addWidget(columns_title)
        columns_header_layout.addWidget(columns_instructions)
        form_layout.addRow(columns_header_widget)

        # ===== ✨ COLONNE CON RADIO BUTTON + BUGFIX =====
        from PyQt6.QtWidgets import QRadioButton, QButtonGroup
        
        self.key_group = QButtonGroup(self)
        self.key_group.setExclusive(True)
        self.column_rows = []

        for i in range(MAX_COLUMNS):
            row_widget = QWidget()
            row_layout = QHBoxLayout(row_widget)
            row_layout.setContentsMargins(0, 0, 0, 0)
            row_layout.setSpacing(10)

            # 🐛 FIX 1: Radio button DISABILITATO di default
            radio = QRadioButton()
            radio.setEnabled(False)  # ✅ Disabilitato finché il campo è vuoto
            radio.setStyleSheet("""
                QRadioButton::indicator {
                    width: 20px;
                    height: 20px;
                }
                QRadioButton::indicator:checked {
                    background-color: #e74c3c;
                    border: 2px solid #c0392b;
                    border-radius: 10px;
                }
                QRadioButton::indicator:disabled {
                    background-color: #e9ecef;
                    border: 2px solid #ced4da;
                }
            """)
            radio.setToolTip("Seleziona questa colonna come chiave")
            self.key_group.addButton(radio, i)

            # Campo testo
            edit = QLineEdit()
            edit.setPlaceholderText(f"Colonna {i+1} (Opzionale)")
            edit.setStyleSheet("""
                QLineEdit {
                    padding: 10px;
                    border: 2px solid #ced4da;
                    border-radius: 4px;
                    font-size: 12px;
                }
                QLineEdit:focus {
                    border: 2px solid #3498db;
                }
            """)
            edit.textChanged.connect(self.clear_error)
            
            # 🐛 FIX 1: Abilita/disabilita radio in base al contenuto
            edit.textChanged.connect(
                lambda text, r=radio: r.setEnabled(bool(text.strip()))
            )

            # Effetto visivo quando selezionato
            def make_highlight_callback(edit_widget, radio_widget):
                def on_toggle(checked):
                    if checked:
                        edit_widget.setStyleSheet("""
                            QLineEdit {
                                padding: 10px;
                                border: 3px solid #e74c3c;
                                border-radius: 4px;
                                font-size: 12px;
                                background-color: #fff3cd;
                                font-weight: bold;
                            }
                            QLineEdit:focus {
                                border: 3px solid #c0392b;
                            }
                        """)
                    else:
                        edit_widget.setStyleSheet("""
                            QLineEdit {
                                padding: 10px;
                                border: 2px solid #ced4da;
                                border-radius: 4px;
                                font-size: 12px;
                            }
                            QLineEdit:focus {
                                border: 2px solid #3498db;
                            }
                        """)
                return on_toggle
            
            radio.toggled.connect(make_highlight_callback(edit, radio))

            row_layout.addWidget(radio)
            row_layout.addWidget(edit)

            self.column_rows.append((radio, edit))
            
            label = QLabel(f"Colonna {i+1}:")
            label.setStyleSheet("color: #6c757d; font-size: 12px;")
            form_layout.addRow(label, row_widget)

        scroll.setWidget(form_container)
        right_layout.addWidget(scroll)

        # Azioni Editor
        editor_actions = QHBoxLayout()
        editor_actions.setSpacing(10)
        
        self.btn_set_default = QPushButton("⭐ Imposta Predefinito")
        self.btn_set_default.setStyleSheet("""
            QPushButton {
                background-color: #f39c12;
                color: white;
                padding: 12px 20px;
                font-weight: bold;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #e67e22;
            }
            QPushButton:disabled {
                background-color: #bdc3c7;
            }
        """)
        self.btn_set_default.clicked.connect(self.action_set_default)
        
        self.btn_save_changes = QPushButton("💾 Salva modifiche")
        self.btn_save_changes.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                color: white;
                padding: 12px 20px;
                font-weight: bold;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
        """)
        self.btn_save_changes.clicked.connect(self.action_save_preset)

        editor_actions.addWidget(self.btn_set_default)
        editor_actions.addStretch()
        editor_actions.addWidget(self.btn_save_changes)
        right_layout.addLayout(editor_actions)

        # Footer
        right_layout.addStretch()
        
        footer_separator = QLabel()
        footer_separator.setStyleSheet("""
            background-color: #dee2e6; 
            min-height: 1px; 
            max-height: 1px;
        """)
        right_layout.addWidget(footer_separator)

        final_buttons = QHBoxLayout()
        final_buttons.setSpacing(15)
        
        btn_cancel = QPushButton("✕ Chiudi")
        btn_cancel.setFixedWidth(120)
        btn_cancel.setFixedHeight(40)
        btn_cancel.setStyleSheet("""
            QPushButton {
                background-color: #95a5a6;
                color: white;
                font-weight: bold;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #7f8c8d;
            }
        """)
        btn_cancel.clicked.connect(self.reject)
        
        self.btn_use = QPushButton("🚀 USA QUESTO PRESET")
        self.btn_use.setFixedHeight(45)
        self.btn_use.setStyleSheet("""
            QPushButton {
                background-color: #27ae60;
                color: white;
                padding: 0px 30px;
                font-size: 15px;
                font-weight: bold;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #229954;
            }
            QPushButton:disabled {
                background-color: #bdc3c7;
            }
        """)
        self.btn_use.clicked.connect(self.action_use_selected)
        
        final_buttons.addWidget(btn_cancel)
        final_buttons.addStretch()
        final_buttons.addWidget(self.btn_use)
        right_layout.addLayout(final_buttons)

        main_layout.addWidget(right_widget, 2)
        
        
    # ============================================================
    # GESTIONE LISTA E SELEZIONE
    # ============================================================

    def load_preset_list(self):
        """Popola la lista dei preset con icone di stato"""
        self.list_presets.clear()
        
        # Ottieni l'ordine salvato dal parent
        if hasattr(self.parent(), 'preset_order'):
            saved_order = self.parent().preset_order
        else:
            saved_order = []
        
        # Crea lista ordinata: prima quelli nell'ordine salvato, poi gli altri alfabeticamente
        ordered_names = []
        remaining_names = set(self.presets.keys())
        
        # Aggiungi preset nell'ordine salvato (se esistono ancora)
        for name in saved_order:
            if name in self.presets:
                ordered_names.append(name)
                remaining_names.discard(name)
        
        # Aggiungi preset rimanenti in ordine alfabetico
        ordered_names.extend(sorted(remaining_names))
        
        for name in ordered_names:
            display_name = name
            if name == self.default_preset:
                display_name = f"⭐ {display_name}"
            if name in self.protected_presets:
                display_name = f"🔒 {display_name}"
            
            item = QListWidgetItem(display_name)
            item.setData(Qt.ItemDataRole.UserRole, name)
            self.list_presets.addItem(item)
        
        total = len(ordered_names)
        protected = len([n for n in ordered_names if n in self.protected_presets])
        custom = total - protected
        self.label_preset_count.setText(
            f"Totale: {total} preset ({custom} personalizzati, {protected} di sistema)"
        )

    def select_preset_in_list(self, name):
        """Sincronizza selezione grafica con nome logico"""
        for i in range(self.list_presets.count()):
            item = self.list_presets.item(i)
            if item.data(Qt.ItemDataRole.UserRole) == name:
                self.list_presets.setCurrentRow(i)
                break

    def on_selection_changed(self, current_item, previous_item):
        """Carica preset usando radio button"""
        if not current_item:
            return
        
        clean_name = current_item.data(Qt.ItemDataRole.UserRole)
        self.current_selected_name = clean_name
        
        preset_data = self.presets.get(clean_name)
        if not preset_data:
            return

        # Esci dalla modalità "nuovo preset"
        self.is_new_preset_mode = False
        self.update_editor_mode()

        # Carica nome preset
        self.input_name.setText(clean_name)
        
        # 🐛 FIX 2: Reset ESPLICITO di tutti i radio button prima di caricare
        for radio, _ in self.column_rows:
            radio.setChecked(False)
        
        # Carica colonne e seleziona chiave
        cols = preset_data['columns']
        key = preset_data['key_column']

        for i, (radio, edit) in enumerate(self.column_rows):
            if i < len(cols):
                edit.setText(cols[i])
                radio.setEnabled(True)  # ← AGGIUNGI: Abilita se c'è testo
                # Seleziona radio se è la colonna chiave
                if cols[i] == key:
                    radio.setChecked(True)
            else:
                edit.clear()
                radio.setEnabled(False)  # ← AGGIUNGI: Disabilita se vuoto
        
        # Aggiorna info selezione
        is_default = (clean_name == self.default_preset)
        is_protected = clean_name in self.protected_presets
        
        info_parts = []
        if is_default:
            info_parts.append("⭐ Preset predefinito")
        if is_protected:
            info_parts.append("🔒 Preset protetto")
        
        if info_parts:
            self.label_selection_info.setText(" • ".join(info_parts))
        else:
            self.label_selection_info.setText("Preset personalizzato")
        
        # Gestione stati
        self.btn_delete.setEnabled(not is_protected)
        self.input_name.setEnabled(not is_protected)
        
        # Gestione pulsanti ordinamento
        current_row = self.list_presets.currentRow()
        self.btn_move_up.setEnabled(current_row > 0)
        self.btn_move_down.setEnabled(current_row < self.list_presets.count() - 1)
        
        self.clear_error()

    # ============================================================
    # GESTIONE MODALITÀ NUOVO/MODIFICA
    # ============================================================

    def update_editor_mode(self):
        """Aggiorna UI in base alla modalità"""
        if self.is_new_preset_mode:
            self.editor_header.setText("🆕 Nuovo Preset")
            self.editor_header.setStyleSheet("""
                font-size: 16px;
                font-weight: bold;
                color: #27ae60;
                padding: 10px;
                border-bottom: 3px solid #27ae60;
            """)
            self.btn_save_changes.setText("➕ Crea preset")
            self.btn_set_default.setEnabled(False)
            self.btn_use.setEnabled(False)
            self.label_selection_info.setText(
                "✏️ Compila le colonne e clicca sul radio button per selezionare la chiave"
            )
        else:
            self.editor_header.setText("✏️ Modifica Configurazione")
            self.editor_header.setStyleSheet("""
                font-size: 16px;
                font-weight: bold;
                color: #2c3e50;
                padding: 10px;
                border-bottom: 2px solid #3498db;
            """)
            self.btn_save_changes.setText("💾 Salva modifiche")
            self.btn_set_default.setEnabled(True)
            self.btn_use.setEnabled(True)

    # ============================================================
    # GESTIONE ERRORI INLINE
    # ============================================================

    def show_error(self, message):
        """Mostra errore inline"""
        self.label_error.setText(f"⚠️ {message}")
        self.label_error.show()

    def clear_error(self):
        """Nascondi errore"""
        self.label_error.hide()

    # ============================================================
    # AZIONI PRINCIPALI
    # ============================================================

    def prepare_new_preset(self):
        """Pulisce form per nuovo preset"""
        self.list_presets.clearSelection()
        self.current_selected_name = None
        self.is_new_preset_mode = True
        
        self.input_name.clear()
        
        # Pulisci tutte le colonne e deseleziona radio
        for radio, edit in self.column_rows:
            radio.setChecked(False)
            edit.clear()
        
        self.input_name.setEnabled(True)
        self.input_name.setFocus()
        
        self.update_editor_mode()
        self.clear_error()

    def action_save_preset(self):
        """✨ SALVA PRESET CON BUGFIX VALIDAZIONE"""
        name = self.input_name.text().strip()
        
        # Raccogli colonne e chiave
        columns = []
        key_column = None

        for i, (radio, edit) in enumerate(self.column_rows):
            text = edit.text().strip()
            if text:  # Solo colonne compilate
                columns.append(text)
                # 🐛 FIX 3: Controlla che ENTRAMBI siano veri
                if radio.isChecked() and text:
                    key_column = text

        # 🐛 FIX 3: Validazione esplicita della chiave
        if not key_column:
            self.show_error(
                "Seleziona una colonna chiave cliccando sul radio button. "
                "Il radio button si abilita automaticamente quando compili il campo."
            )
            return

        # Validazione completa
        is_valid, error_msg = self.validate_preset_data(name, key_column, columns)
        if not is_valid:
            self.show_error(error_msg)
            return

        # Controllo sovrascrittura
        if name in self.presets and name != self.current_selected_name:
            reply = QMessageBox.question(
                self,
                "⚠️ Preset Esistente",
                f"Il preset '<b>{name}</b>' esiste già.\n\n"
                "Vuoi sovrascriverlo con questa nuova configurazione?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            if reply == QMessageBox.StandardButton.No:
                return

        # ✅ CORREZIONE: Prepara i dati da restituire
        self.result = {
            'action': 'create',
            'preset_name': name,
            'key_column': key_column,
            'columns': columns,
            'set_as_default': False
        }
        
        # ✅ Salva nel dizionario locale (per preview nel dialog)
        self.presets[name] = {
            'key_column': key_column,
            'columns': columns
        }
        
        # Aggiungi all'ordine se è un nuovo preset
        parent = self.parent()
        if name not in parent.preset_order:
            parent.preset_order.append(name)
        
        # Ricarica lista e seleziona il preset salvato
        self.load_preset_list()
        self.select_preset_in_list(name)
        
        # Esci dalla modalità nuovo
        self.is_new_preset_mode = False
        self.current_selected_name = name
        self.update_editor_mode()
        
        # Feedback
        self.label_selection_info.setText(
            f"✅ Preset '{name}' salvato con successo! "
            f"Colonna chiave: {key_column}\n"
            f"Premi '🚀 USA QUESTO PRESET' per applicarlo"
        )

    def action_set_default(self):
        """Imposta preset come predefinito"""
        if not self.current_selected_name:
            self.show_error("Seleziona un preset prima di impostarlo come predefinito")
            return
        
        if self.current_selected_name == self.default_preset:
            QMessageBox.information(
                self,
                "Già Predefinito",
                f"Il preset '<b>{self.current_selected_name}</b>' è già impostato come predefinito!"
            )
            return
        
        self.result = {
            'action': 'set_default', 
            'preset_name': self.current_selected_name
        }
        self.accept()

    def action_use_selected(self):
        """Usa il preset selezionato"""
        if not self.current_selected_name:
            self.show_error("Seleziona un preset prima di usarlo")
            return
        
        # ✅ FIX: Se abbiamo appena creato questo preset, mantieni action='create'
        if (self.result and 
            self.result.get('action') == 'create' and 
            self.result.get('preset_name') == self.current_selected_name):
            # Usa il result esistente che contiene tutti i dati
            self.accept()
        else:
            # Preset esistente, usa action='select'
            self.result = {
                'action': 'select', 
                'preset_name': self.current_selected_name
            }
            self.accept()

    def delete_selected_preset(self):
        """Elimina preset selezionato"""
        if not self.current_selected_name:
            self.show_error("Seleziona un preset prima di eliminarlo")
            return
        
        if self.current_selected_name in self.protected_presets:
            self.show_error(
                f"Il preset '{self.current_selected_name}' è protetto e non può essere eliminato"
            )
            return
        
        confirm = QMessageBox.question(
            self, 
            "🗑️ Conferma Eliminazione", 
            f"Sei sicuro di voler eliminare definitivamente il preset "
            f"'<b>{self.current_selected_name}</b>'?\n\n"
            "Questa azione non può essere annullata.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )
        
        if confirm == QMessageBox.StandardButton.Yes:
            preset_to_delete = self.current_selected_name
            
            # ✅ ELIMINA LOCALMENTE (non chiudere il dialog)
            # Rimuovi dal dizionario locale
            del self.presets[preset_to_delete]
            
            # Rimuovi dall'ordine
            parent = self.parent()
            if preset_to_delete in parent.preset_order:
                parent.preset_order.remove(preset_to_delete)
            
            # Se era il preset predefinito, resettalo
            if self.default_preset == preset_to_delete:
                self.default_preset = None
            
            # ✅ SALVA SUBITO (così il parent è sincronizzato)
            parent.presets = self.presets.copy()
            parent.default_preset = self.default_preset
            parent.save_presets()
            
            # ✅ RICARICA LA LISTA E SELEZIONA IL PRIMO DISPONIBILE
            self.load_preset_list()
            if self.presets:
                first_preset_name = list(self.presets.keys())[0]
                self.select_preset_in_list(first_preset_name)
            else:
                # Nessun preset rimasto (teoricamente impossibile, ci sono quelli protetti)
                self.prepare_new_preset()
            
            # ✅ FEEDBACK VISIVO
            self.label_selection_info.setText(
                f"✅ Preset '{preset_to_delete}' eliminato con successo"
            )
            
            
    def move_preset_up(self):
        """Sposta il preset selezionato verso l'alto nella lista"""
        current_row = self.list_presets.currentRow()
        
        if current_row <= 0:
            return  # Già in cima
        
        current_item = self.list_presets.currentItem()
        preset_name = current_item.data(Qt.ItemDataRole.UserRole)
        
        # Aggiorna l'ordine nel parent
        self._update_preset_order()
        parent = self.parent()
        
        if preset_name in parent.preset_order:
            current_index = parent.preset_order.index(preset_name)
            if current_index > 0:
                # Scambia con il precedente
                parent.preset_order[current_index], parent.preset_order[current_index - 1] = \
                    parent.preset_order[current_index - 1], parent.preset_order[current_index]
        
        # Ricarica la lista e mantieni la selezione
        self.load_preset_list()
        self.list_presets.setCurrentRow(current_row - 1)
        
        # Salva immediatamente
        parent.save_presets()
        
        self.label_selection_info.setText("✅ Preset spostato verso l'alto")

    def move_preset_down(self):
        """Sposta il preset selezionato verso il basso nella lista"""
        current_row = self.list_presets.currentRow()
        
        if current_row >= self.list_presets.count() - 1:
            return  # Già in fondo
        
        current_item = self.list_presets.currentItem()
        preset_name = current_item.data(Qt.ItemDataRole.UserRole)
        
        # Aggiorna l'ordine nel parent
        self._update_preset_order()
        parent = self.parent()
        
        if preset_name in parent.preset_order:
            current_index = parent.preset_order.index(preset_name)
            if current_index < len(parent.preset_order) - 1:
                # Scambia con il successivo
                parent.preset_order[current_index], parent.preset_order[current_index + 1] = \
                    parent.preset_order[current_index + 1], parent.preset_order[current_index]
        
        # Ricarica la lista e mantieni la selezione
        self.load_preset_list()
        self.list_presets.setCurrentRow(current_row + 1)
        
        # Salva immediatamente
        parent.save_presets()
        
        self.label_selection_info.setText("✅ Preset spostato verso il basso")

    def _update_preset_order(self):
        """Sincronizza l'ordine corrente della lista con preset_order del parent"""
        parent = self.parent()
        current_order = []
        
        for i in range(self.list_presets.count()):
            item = self.list_presets.item(i)
            preset_name = item.data(Qt.ItemDataRole.UserRole)
            current_order.append(preset_name)
        
        parent.preset_order = current_order

    def get_result(self):
        """Restituisce il risultato della dialog"""
        return self.result   
        
        
        
               

class ComparisonEngine:
    """Motore di confronto logico separato dalla UI"""
    
    @staticmethod
    def count_duplicates(df: pd.DataFrame, key_column: str = 'KEY_NORMALIZED') -> tuple[int, int]:
        """
        Conta duplicati in un DataFrame.
        
        Args:
            df: DataFrame con colonna chiave
            key_column: Nome colonna chiave (default: 'KEY_NORMALIZED')
        
        Returns:
            tuple: (unique_keys_duplicated, total_rows_duplicated)
        """
        if df.empty or key_column not in df.columns:
            return 0, 0
        
        # Filtra righe con chiave valida
        valid_mask = df[key_column].notna() & (df[key_column].astype(str).str.strip() != "")
        df_valid = df[valid_mask]
        
        if df_valid.empty:
            return 0, 0
        
        duplicates_mask = df_valid[key_column].duplicated(keep=False)
        unique_keys = df_valid[duplicates_mask][key_column].nunique()
        total_rows = duplicates_mask.sum()
        
        return unique_keys, total_rows
    
    @staticmethod
    def compare(df1, df2, key_column='KEY_NORMALIZED'):
        """
        Confronta due DataFrame basandosi sulla colonna chiave normalizzata.
        
        Args:
            df1: DataFrame principale (worklist)
            df2: DataFrame di confronto
            key_column: Nome della colonna normalizzata (default: 'KEY_NORMALIZED')
        
        Returns:
            tuple: (matches_df, mismatches_df, stats_dict)
        """
        # Estrai le chiavi normalizzate
        keys_df2 = set(df2[key_column].dropna())
        
        # Identifica corrispondenze e mancanti
        matches_mask = df1[key_column].isin(keys_df2)
        matches = df1[matches_mask].copy()
        mismatches = df1[~matches_mask].copy()
        
        # Calcola statistiche usando il metodo centralizzato
        duplicates_df1, _ = ComparisonEngine.count_duplicates(df1, key_column)
        duplicates_df2, _ = ComparisonEngine.count_duplicates(df2, key_column)
        
        stats = {
            'total': len(df1),
            'matches': len(matches),
            'mismatches': len(mismatches),
            'duplicates_df1': duplicates_df1,
            'duplicates_df2': duplicates_df2
        }
        
        return matches, mismatches, stats


class TableManager:
    """Gestore centralizzato per la configurazione e manipolazione delle tabelle"""
    
    def __init__(self, columns=None):
        self.columns = columns if columns is not None else []
        self._sort_column = None
        self._sort_order = None
    
    def configure_table_widget(self, table):
        """Configurazione centralizzata dello stile e del comportamento delle tabelle"""
        table.setColumnCount(len(self.columns))
        table.setHorizontalHeaderLabels(self.columns)
        table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        table.setAlternatingRowColors(True)
        table.setSortingEnabled(False)
        table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        table.setSelectionMode(QTableWidget.SelectionMode.ExtendedSelection)
        
        # Imposta altezza minima righe per evitare tagli
        table.verticalHeader().setDefaultSectionSize(35)
        table.verticalHeader().setMinimumSectionSize(35)
    
    def enable_manual_sorting(self, table):
        """Abilita sorting manuale con indicatori Unicode visibili"""
        header = table.horizontalHeader()
        header.setSortIndicatorShown(False)
        header.setSectionsClickable(True)
        
        # Salva il riferimento all'handler per poterlo disconnettere dopo
        if not hasattr(self, '_sort_handlers'):
            self._sort_handlers = {}
        
        # Disconnetti l'handler precedente di QUESTA tabella specifica
        if table in self._sort_handlers:
            try:
                header.sectionClicked.disconnect(self._sort_handlers[table])
            except TypeError:
                pass
        
        def on_header_clicked(logical_index):
            # Determina la direzione dell'ordinamento
            if self._sort_column == logical_index:
                self._sort_order = Qt.SortOrder.DescendingOrder if self._sort_order == Qt.SortOrder.AscendingOrder else Qt.SortOrder.AscendingOrder
            else:
                self._sort_column = logical_index
                self._sort_order = Qt.SortOrder.AscendingOrder
            
            # Ordina i dati
            table.setSortingEnabled(True)
            table.sortItems(logical_index, self._sort_order)
            table.setSortingEnabled(False)
            
            # Aggiorna le intestazioni con frecce Unicode
            for col in range(table.columnCount()):
                # ✅ CORREZIONE: Verifica che l'indice sia valido per le colonne correnti
                if col < len(self.columns):
                    original_text = self.columns[col]
                    if col == logical_index:
                        arrow = " ▲" if self._sort_order == Qt.SortOrder.AscendingOrder else " ▼"
                        table.horizontalHeaderItem(col).setText(original_text + arrow)
                    else:
                        table.horizontalHeaderItem(col).setText(original_text)
        
        # Salva il riferimento per disconnessione futura
        self._sort_handlers[table] = on_header_clicked
        header.sectionClicked.connect(on_header_clicked)

    def reset_sorting(self, table):
        """Rimuove gli indicatori di sorting dalla tabella"""
        self._sort_column = None
        self._sort_order = None
        
        # Ripristina le intestazioni originali senza frecce
        for col in range(table.columnCount()):
            if col < len(self.columns):
                table.horizontalHeaderItem(col).setText(self.columns[col])

class CodeComparator(QMainWindow):
    """Applicazione principale per confrontare liste basate su colonne chiave"""
    
    @property
    def columns(self):
        """Restituisce le colonne correnti"""
        return self._columns
    
    @columns.setter
    def columns(self, nuovo_valore):
        """Aggiorna colonne e sincronizza con table_manager"""
        if not isinstance(nuovo_valore, list):
            raise ValueError("columns deve essere una lista")
        
        self._columns = nuovo_valore
        
        # Sincronizza con table_manager se esiste
        if hasattr(self, 'table_manager'):
            self.table_manager.columns = nuovo_valore
    
    
    def __init__(self):
        super().__init__()
        QApplication.setStyle('Fusion')
        
        self.setWindowTitle("Code Comparator - Confronto Liste")
        self.resize(1200, 900)
        self.setMinimumSize(800, 600)  # ✅ AGGIUNGI QUESTA RIGA

        # ============================================================
        # CONFIGURAZIONE PRESET
        # ============================================================
        self.presets = {
            "ACQUISTI (ISBN, Titolo, Autore, Editore, Anno, Prezzo, Note)": {
                "key_column": "ISBN",
                "columns": ["ISBN", "TITOLO", "AUTORE", "EDITORE", "ANNO", "PREZZO", "NOTE"]
            },
            "ACQUISTI 2 (ISBN, Titolo)": {
                "key_column": "ISBN",
                "columns": ["ISBN", "TITOLO"]
            },
            "STAMPA REGISTRI 1 (Inventario, Sezione, Collocazione, Specificazione, Sequenza, Descrizione ISBD, Legami)": {
                "key_column": "Inventario",
                "columns": ["Inventario", "Sezione", "Collocazione", "Specificazione", 
                           "Sequenza", "Descrizione ISBD", "Legami"]
            },
            "STAMPA REGISTRI 2 (Inventario, Descrizione ISBD)": {
                "key_column": "Inventario",
                "columns": ["Inventario", "Descrizione ISBD"]
            }
        }
        
        # *** NUOVO: Set di preset protetti (non eliminabili) ***
        self.protected_presets = set(self.presets.keys())
        
       
        # --- File di configurazione nella cartella dello script ---
        try:
            base_dir = Path(__file__).resolve().parent
        except NameError:
            # Fallback se __file__ non è definito (es. in alcuni contesti di esecuzione)
            base_dir = Path.cwd()
        self.config_file = str(base_dir / "presets.json")
        
        # --- Variabile per preset predefinito ---
        self.default_preset = None
        
        self.preset_order = []  # ← CREA la variabile PRIMA di caricare
        
        # --- Carica preset salvati ---
        self.load_presets()
        
        # --- Usa il preset predefinito se esiste, altrimenti usa ACQUISTI ---
        if self.default_preset and self.default_preset in self.presets:
            self.current_preset = self.default_preset
        else:
            self.current_preset = "ACQUISTI (ISBN, Titolo, Autore, Editore, Anno, Prezzo, Note)"
        
        self.key_column = self.presets[self.current_preset]["key_column"]
        self._columns = self.presets[self.current_preset]["columns"].copy()
        
        # ============================================================
        # TABLE MANAGER
        # ============================================================
        self.table_manager = TableManager(self._columns)
        
        # ============================================================
        # CENTRALIZZAZIONE: Dizionario unico per gestire le liste
        # ============================================================
        self.lists = {
            1: {
                "df": pd.DataFrame(columns=self.columns),
                "table": None,
                "label": None,
                "duplicates_label": None,  # ✅ AGGIUNGI QUESTO
                "deleted": [],
                "btn_restore": None
            },
            2: {
                "df": pd.DataFrame(columns=self.columns),
                "table": None,
                "label": None,
                "duplicates_label": None,  # ✅ AGGIUNGI QUESTO
                "deleted": [],
                "btn_restore": None
            }
        }
                
        

        # Risultati del confronto
        self.matches_df = None
        self.mismatches_df = None

        self.init_ui()
        self.apply_styles()
    
    
    
    
    # ============================================================
    # PERSISTENZA PRESET
    # ============================================================
    
    def save_presets(self):
        """
        Salva i preset attuali e il preset predefinito in presets.json nella cartella dello script.
        Crea un backup prima di sovrascrivere.
        """
        try:
            # Backup del file esistente
            if os.path.exists(self.config_file):
                backup_file = self.config_file + ".bak"
                try:
                    shutil.copy2(self.config_file, backup_file)
                except Exception as e:
                    logging.warning(f"Impossibile creare backup: {e}")
            
            # Prepara i dati da salvare con preset predefinito e ordine
            data_to_save = {
                "default_preset": self.default_preset,
                "preset_order": self.preset_order if hasattr(self, 'preset_order') else [],
                "presets": self.presets
            }
            
            # Salva i preset
            with open(self.config_file, "w", encoding="utf-8") as f:
                json.dump(data_to_save, f, indent=4, ensure_ascii=False)
            
            logging.info(f"Preset salvati in: {self.config_file}")  
            
        except PermissionError as e:
            QMessageBox.critical(
                self,
                "⌠Errore Critico - Permessi",
                f"Impossibile scrivere su:\n{self.config_file}\n\n{str(e)}"
            )
            logging.exception(f"PermissionError salvando preset: {e}")
        except Exception as e:
            QMessageBox.critical(
                self, 
                "⌠Errore Critico - Salvataggio",
                f"Impossibile salvare i preset:\n\n{str(e)}\n\n"
                "I preset saranno persi alla chiusura."
            )
            logging.exception(f"Errore nel salvataggio dei preset: {e}")

    def load_presets(self):
        """
        Carica i preset e il preset predefinito da presets.json se esiste.
        Valida i dati caricati per evitare corruzione.
        """
        if not os.path.exists(self.config_file):
            logging.info("Nessun file presets.json trovato. Uso preset predefiniti.")  
            return
        
        try:
            with open(self.config_file, "r", encoding="utf-8") as f:
                loaded_data = json.load(f)
            
            # Gestisci retrocompatibilità con vecchio formato (solo preset senza struttura)
            if isinstance(loaded_data, dict):
                # Nuovo formato con "default_preset" e "presets"
                if "presets" in loaded_data:
                    loaded_presets = loaded_data.get("presets", {})
                    self.default_preset = loaded_data.get("default_preset", None)
                    self.preset_order = loaded_data.get("preset_order", [])
                # Vecchio formato (solo preset diretti)
                else:
                    loaded_presets = loaded_data
                    self.default_preset = None
                    self.preset_order = []
            else:
                loaded_presets = {}
                self.default_preset = None
                self.preset_order = []
            
            # Validazione: ogni preset deve avere "key_column" e "columns"
            valid_presets = {}
            for preset_name, preset_data in loaded_presets.items():
                if self._validate_preset(preset_name, preset_data):
                    valid_presets[preset_name] = preset_data
                else:
                    logging.warning(f"Preset '{preset_name}' non valido, ignorato.")
            
            # Unisci con i preset predefiniti (i custom sovrascrivono i default se hanno lo stesso nome)
            self.presets.update(valid_presets)
            
            # Verifica che il preset predefinito esista ancora
            if self.default_preset and self.default_preset not in self.presets:
                logging.warning(f"Preset predefinito '{self.default_preset}' non trovato. Resettato.")
                self.default_preset = None
            
            logging.info(f"Caricati {len(valid_presets)} preset da: {self.config_file}")
            
            
            
        except json.JSONDecodeError as e:
            QMessageBox.critical(
                self,
                "⌠Errore Critico - File Corrotto",
                f"Il file presets.json è corrotto:\n\n{str(e)}\n\n"
                "Tentativo di ripristino dal backup..."
            )
            logging.exception(f"Errore parsing JSON: {e}")
            self._restore_backup()
            
        except Exception as e:
            logging.exception(f"Errore nel caricamento dei preset: {e}")

    def _validate_preset(self, preset_name, preset_data):
        """
        Valida la struttura di un preset usando la validazione unificata.
        
        Args:
            preset_name: Nome del preset
            preset_data: Dizionario con "key_column" e "columns"
        
        Returns:
            True se valido, False altrimenti
        """
        if not isinstance(preset_data, dict):
            return False
        
        if "key_column" not in preset_data or "columns" not in preset_data:
            return False
        
        # Usa la validazione unificata di SettingsDialog
        is_valid, _ = SettingsDialog.validate_preset_data(
            preset_name,
            preset_data["key_column"],
            preset_data["columns"]
        )
        
        return is_valid

    def _restore_backup(self):
        """Tenta di ripristinare il backup in caso di file corrotto"""
        backup_file = self.config_file + ".bak"
        
        if os.path.exists(backup_file):
            try:
                shutil.copy2(backup_file, self.config_file)
                logging.info(f"Backup ripristinato da: {backup_file}")
                
                # Riprova a caricare
                self.load_presets()
            except Exception as e:
                QMessageBox.critical(
                    self,
                    "⌠Errore Critico - Ripristino Backup",
                    f"Impossibile ripristinare il backup:\n\n{str(e)}\n\n"
                    "Verifica manualmente il file presets.json"
                )
                logging.exception(f"Impossibile ripristinare il backup: {e}")
    
    # ============================================================
    # GESTIONE EVENTI TASTIERA
    # ============================================================
    
    def _get_active_list_id(self):
        """Restituisce l'ID della lista con focus, o None"""
        focused = QApplication.focusWidget()
        for list_id, data in self.lists.items():
            if focused == data["table"] or data["table"].isAncestorOf(focused):
                return list_id
        return None

    def keyPressEvent(self, event):
        """Gestisce CTRL+Z per ripristinare righe eliminate e CTRL+C per copiare"""
        list_id = self._get_active_list_id()
        
        if list_id is None:
            super().keyPressEvent(event)
            return
        
        # CTRL+Z: Ripristina righe eliminate
        if event.matches(QKeySequence.StandardKey.Undo):
            if self.lists[list_id]["deleted"]:
                self.restore_rows(list_id)
            event.accept()
        
        # CTRL+C: Copia righe selezionate
        elif event.matches(QKeySequence.StandardKey.Copy):
            if self.lists[list_id]["table"].selectedItems():
                self.copy_selected_rows(list_id)
            event.accept()
        
        else:
            super().keyPressEvent(event)
    

    # ============================================================
    # INTERFACCIA UTENTE
    # ============================================================

    def init_ui(self):
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        main_layout.setSpacing(15)
        main_layout.setContentsMargins(15, 10, 15, 15)

        # --- HEADER: TITOLO E IMPOSTAZIONI ---
        header_layout = QHBoxLayout()
        
        title_label = QLabel("📚 Code Comparator - Confronto Liste")
        title_label.setStyleSheet("font-size: 20px; font-weight: bold; color: #2c3e50; padding: 5px;")
        title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        
        btn_settings = QPushButton("⚙️ Impostazioni")
        btn_settings.setMinimumSize(140, 40)  # Permette espansione
        btn_settings.clicked.connect(self.open_settings)
        btn_settings.setObjectName("btn_settings")
        
        header_layout.addStretch()
        header_layout.addWidget(title_label)
        header_layout.addStretch()
        header_layout.addWidget(btn_settings)
        
        main_layout.addLayout(header_layout)

        # --- SEZIONE INPUT: TABELLE LISTE ---
        input_layout = QHBoxLayout()
        
        self.lists[1]["table"] = self.create_table_group(input_layout, "📋 Lista 1 (Worklist)", 1)
        self.lists[2]["table"] = self.create_table_group(input_layout, "📂 Lista 2 (Confronto)", 2)
        
        main_layout.addLayout(input_layout)

        # --- BOTTONI AZIONI PRINCIPALI ---
        buttons_layout = QHBoxLayout()
        buttons_layout.setSpacing(10)

        # Pulsante Inverti Liste
        self.btn_swap = QPushButton("🔄 INVERTI LISTE")
        self.btn_swap.setFixedHeight(60)
        self.btn_swap.clicked.connect(self.swap_lists)
        self.btn_swap.setToolTip("Scambia Lista 1 con Lista 2")

        self.btn_compare = QPushButton("🔍 CONFRONTA LISTE")
        self.btn_compare.setFixedHeight(60)
        self.btn_compare.clicked.connect(self.compare_data)

        self.btn_export_matches = QPushButton("💾 ESPORTA CORRISPONDENZE")
        self.btn_export_matches.setFixedHeight(60)
        self.btn_export_matches.clicked.connect(lambda: self.export_results("matches"))
        self.btn_export_matches.setEnabled(False)

        self.btn_export_mismatches = QPushButton("💾 ESPORTA MANCANTI")
        self.btn_export_mismatches.setFixedHeight(60)
        self.btn_export_mismatches.clicked.connect(lambda: self.export_results("mismatches"))
        self.btn_export_mismatches.setEnabled(False)

        self.btn_clear_results = QPushButton("🗑️ CANCELLA RISULTATI")
        self.btn_clear_results.setFixedHeight(60)
        self.btn_clear_results.clicked.connect(self.clear_results)
        self.btn_clear_results.setEnabled(False)

        buttons_layout.addWidget(self.btn_swap)
        buttons_layout.addWidget(self.btn_compare)
        buttons_layout.addWidget(self.btn_export_matches)
        buttons_layout.addWidget(self.btn_export_mismatches)
        buttons_layout.addWidget(self.btn_clear_results)
        
        main_layout.addLayout(buttons_layout)

        # --- SEZIONE RISULTATI (TABS) ---
        self.tabs_results = QTabWidget()
        
        self.res_matches = self._create_result_table()
        self.res_mismatches = self._create_result_table()
        
        
        
        self.tabs_results.addTab(self.res_matches, "✓ Corrispondenze (0)")
        self.tabs_results.addTab(self.res_mismatches, "✗ Mancanti (0)")
        
        main_layout.addWidget(self.tabs_results)

        # --- BARRA INFORMAZIONI ---
        self.info_label = QLabel("Pronto per il confronto. Incolla i dati da Excel/Google Sheets.")
        self.info_label.setStyleSheet("padding: 8px; background-color: #ecf0f1; border-radius: 5px;")
        main_layout.addWidget(self.info_label)

    def create_table_group(self, layout, title, list_id):
        """
        Crea un gruppo tabella con controlli per una lista specifica.
        CENTRALIZZATO: Usa list_id per identificare la lista.
        """
        container = QWidget()
        v_layout = QVBoxLayout(container)
        v_layout.setSpacing(10)
        
        # Header con titolo e contatore
        header_layout = QHBoxLayout()
        label = QLabel(title)
        label.setStyleSheet("font-weight: bold; font-size: 14px; color: #34495e;")

        count_label = QLabel("(0 righe)")
        count_label.setStyleSheet("color: #7f8c8d; font-size: 12px;")

        # Label duplicati cliccabile
        duplicates_label = QLabel("")
        duplicates_label.setStyleSheet(
            "color: #e74c3c; font-size: 12px; text-decoration: underline;"  # ✅ Rimosso "cursor: pointer"
        )
        duplicates_label.setCursor(Qt.CursorShape.PointingHandCursor)  # ✅ Questo è il modo corretto in Qt
        duplicates_label.mousePressEvent = lambda event: self.show_duplicates_dialog(list_id)

        header_layout.addWidget(label)
        header_layout.addWidget(count_label)
        header_layout.addWidget(duplicates_label)
        header_layout.addStretch()
        
        # Bottoni azione (4 pulsanti) - CENTRALIZZATI con lambda
        btn_layout = QHBoxLayout()
        
        btn_paste = QPushButton("📋 Incolla da Excel")
        btn_paste.clicked.connect(lambda: self.paste_data(list_id))
        btn_paste.setFixedHeight(35)
        
        btn_clear = QPushButton("🗑️ Cancella lista")
        btn_clear.clicked.connect(lambda: self.clear_data(list_id))
        btn_clear.setFixedHeight(35)
        
        btn_delete = QPushButton("❌ Elimina righe")
        btn_delete.clicked.connect(lambda: self.delete_rows(list_id))
        btn_delete.setFixedHeight(35)
        
        btn_restore = QPushButton("↩️ Ripristina righe")
        btn_restore.clicked.connect(lambda: self.restore_rows(list_id))
        btn_restore.setFixedHeight(35)
        btn_restore.setEnabled(False)
        
        btn_layout.addWidget(btn_paste)
        btn_layout.addWidget(btn_clear)
        btn_layout.addWidget(btn_delete)
        btn_layout.addWidget(btn_restore)
        
        # Tabella con menu contestuale
        table = QTableWidget(0, len(self.columns))
        self.table_manager.configure_table_widget(table)
        self.table_manager.enable_manual_sorting(table)
        table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        
        table.customContextMenuRequested.connect(lambda pos: self.show_context_menu(pos, list_id))
        
        v_layout.addLayout(header_layout)
        v_layout.addLayout(btn_layout)
        v_layout.addWidget(table)
        
        layout.addWidget(container)
        
        # Salva riferimenti nel dizionario centralizzato
        store = self.lists[list_id]
        store["table"] = table
        store["label"] = count_label
        store["duplicates_label"] = duplicates_label
        store["btn_restore"] = btn_restore

        return table

    

    def _create_result_table(self):
        """Crea e configura una tabella risultato"""
        table = QTableWidget()
        self.table_manager.configure_table_widget(table)
        self.table_manager.enable_manual_sorting(table)
        return table


    # ============================================================
    # METODI UNIFICATI PER GESTIONE LISTE
    # ============================================================
    
    def _update_counters(self, list_id):
        """Aggiorna contatori righe e duplicati per una lista"""
        store = self.lists[list_id]
        df = store["df"]
        
        # Aggiorna sempre il contatore righe
        store["label"].setText(f"({len(df)} righe)")
        
        # Salta il calcolo duplicati se vuoto
        if df.empty:
            store["duplicates_label"].setText("")
            return
        
        # Filtra righe con chiave valida
        df_with_keys = self._get_df_with_valid_keys(df)
        
        if df_with_keys.empty:
            store["duplicates_label"].setText("")
            return
        
        duplicates_unique, duplicates_count = ComparisonEngine.count_duplicates(df_with_keys)
        
        # Aggiorna label duplicati
        if duplicates_count > 0:
            store["duplicates_label"].setText(
                f"⚠️ {duplicates_unique} duplicati ({duplicates_count} righe)"
            )
        else:
            store["duplicates_label"].setText("")
    

    def paste_data(self, list_id: int) -> None:
        """Metodo UNIFICATO per incollare dati in qualsiasi lista"""
        new_df = self.get_clipboard_data()
        if new_df is None:
            return
        
        store = self.lists[list_id]
        old_row_count = len(store["df"])
        
        # Aggiungi i nuovi dati (gestisci numerazione progressiva)
        if not store["df"].empty:
            # ✅ Continua la numerazione delle righe originali
            last_row = store["df"]['__ORIGINAL_ROW__'].max()
            new_df['__ORIGINAL_ROW__'] = new_df['__ORIGINAL_ROW__'] + last_row
            store["df"] = pd.concat([store["df"], new_df], ignore_index=True)
        else:
            # ✅ Prima volta: usa il DataFrame così com'è
            store["df"] = new_df
        
        self.update_table_display(store["table"], store["df"], highlight_from=old_row_count)
        self._update_counters(list_id) 
        self.table_manager.reset_sorting(store["table"])

        new_rows = len(new_df)
        if old_row_count > 0:
            self.info_label.setText(f"✅ Lista {list_id}: aggiunte {new_rows} righe (totale: {len(store['df'])})")
            logging.info(f"Lista {list_id}: incollate {new_rows} righe (totale: {len(store['df'])})")
        else:
            self.info_label.setText(f"✅ Lista {list_id} caricata: {len(store['df'])} righe")
            logging.info(f"Lista {list_id}: caricata con {len(store['df'])} righe")

    def clear_data(self, list_id: int) -> None:
        """Cancella tutti i dati di una lista specifica"""
        self._reset_single_list(list_id)  # ✅ USA IL NUOVO HELPER
        self.info_label.setText(f"Lista {list_id} cancellata")

    def delete_rows(self, list_id: int) -> None:
        """Metodo UNIFICATO per eliminare righe selezionate"""
        store = self.lists[list_id]
        table = store["table"]
        df = store["df"]
        
        selected_rows_ui = sorted(set(item.row() for item in table.selectedItems()))
        
        if not selected_rows_ui:
            QMessageBox.warning(self, "Nessuna Selezione", 
                              "Seleziona almeno una riga da eliminare!")
            return
        
        reply = QMessageBox.question(
            self, 
            "Conferma Eliminazione",
            f"Vuoi eliminare {len(selected_rows_ui)} riga/e selezionata/e?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )
        
        if reply != QMessageBox.StandardButton.Yes:
            return
        
        indices_to_delete = []
        deleted_rows_data = []
        
        for row_idx in selected_rows_ui:
            # ✅ CORRETTO: Cerca il primo item disponibile in qualsiasi colonna
            item = None
            for col in range(table.columnCount()):
                item = table.item(row_idx, col)
                if item:
                    break
            
            if not item:
                logging.warning(f"Riga {row_idx}: nessun item trovato in alcuna colonna")
                continue
            
            real_index = item.data(Qt.ItemDataRole.UserRole)
            if real_index is None:
                logging.warning(f"Riga {row_idx}: UserRole non impostato")
                continue
            
            if real_index is not None and isinstance(real_index, int) and 0 <= real_index < len(df):
                indices_to_delete.append(real_index)
                try:
                    deleted_rows_data.append({
                        'index': real_index,
                        'data': df.iloc[real_index].to_dict()
                    })
                except Exception as e:
                    logging.error(f"Errore nel salvataggio riga {real_index}: {e}")
        
        if not indices_to_delete:
            QMessageBox.warning(self, "Errore", 
                "Impossibile identificare le righe da eliminare!\n"
                "Prova a ricaricare i dati.")
            return
        
        store["deleted"].append(deleted_rows_data)
        df_updated = df.drop(indices_to_delete).reset_index(drop=True)
        store["df"] = df_updated
        
        self.update_table_display(store["table"], df_updated)
        self._update_counters(list_id)
        self.table_manager.reset_sorting(store["table"]) 
        store["btn_restore"].setEnabled(True)

        logging.info(f"Lista {list_id}: eliminate {len(indices_to_delete)} righe")
        self.info_label.setText(f"✅ Eliminate {len(indices_to_delete)} righe dalla Lista {list_id}")

    
    def restore_rows(self, list_id: int) -> None:
        """Metodo UNIFICATO per ripristinare righe eliminate"""
        store = self.lists[list_id]
        
        if not store["deleted"]:
            QMessageBox.information(self, "Nessuna Operazione", 
                                  "Non ci sono eliminazioni da ripristinare!")
            return
        
        last_deleted = store["deleted"].pop()
        df = store["df"]
        
        try:
            # Ordina per indice decrescente per evitare shift
            sorted_deleted = sorted(last_deleted, key=lambda x: x['index'], reverse=True)
            
            # ✅ VALIDAZIONE: Verifica che le colonne siano ancora compatibili
            if last_deleted:
                sample_row = last_deleted[0]['data']
                missing_cols = set(sample_row.keys()) - set(df.columns)
                if missing_cols:
                    raise ValueError(
                        f"Impossibile ripristinare: colonne mancanti nel DataFrame corrente: {missing_cols}"
                    )
            
            # Reinserisci le righe nelle posizioni originali
            for item in sorted_deleted:
                original_index = item['index']
                row_data = pd.Series(item['data'])
                
                # Inserisci la riga nella posizione originale
                if original_index <= len(df):
                    df = pd.concat([
                        df.iloc[:original_index],
                        pd.DataFrame([row_data]),
                        df.iloc[original_index:]
                    ], ignore_index=True)
                else:
                    df = pd.concat([df, pd.DataFrame([row_data])], ignore_index=True)
            
            store["df"] = df
            
            # Trova le posizioni delle righe ripristinate per evidenziarle
            restored_positions = sorted([item['index'] for item in last_deleted])
            min_position = min(restored_positions) if restored_positions else 0
            
            self.update_table_display(store["table"], store["df"], highlight_from=min_position)
            self._update_counters(list_id)
            self.table_manager.reset_sorting(store["table"])  # AGGIUNGI QUESTA RIGA

            if not store["deleted"]:
                store["btn_restore"].setEnabled(False)

            self.info_label.setText(f"✅ Ripristinate {len(last_deleted)} righe nella Lista {list_id}")
            
        except Exception as e:
            # Ripristina lo stato precedente in caso di errore
            store["deleted"].append(last_deleted)
            logging.exception(f"Errore nel ripristino righe: {e}")
            QMessageBox.critical(
                self,
                "❌ Errore Ripristino",
                f"Impossibile ripristinare le righe:\n{str(e)}\n\n"
                "Le righe sono state mantenute nello storico eliminazioni."
            )
    
    

    def copy_selected_rows(self, list_id: int) -> None:
        """Copia le righe selezionate negli appunti (formato Excel-compatibile)"""
        store = self.lists[list_id]
        table = store["table"]
        
        selected_rows = sorted(set(item.row() for item in table.selectedItems()))
        
        if not selected_rows:
            QMessageBox.warning(self, "Nessuna Selezione", 
                              "Seleziona almeno una riga da copiare!")
            return
        
        copied_data = []
        num_cols = table.columnCount()  # ✅ Usa il numero di colonne EFFETTIVE della tabella
        for row_idx in selected_rows:
            row_data = []
            for col_idx in range(num_cols):  
                item = table.item(row_idx, col_idx)
                cell_value = item.text() if item else ""
                
                # ✅ SOLUZIONE: Racchiudi tra virgolette se contiene newline o tab
                if '\n' in cell_value or '\t' in cell_value or '"' in cell_value:
                    # Escape virgolette doppie (come fa Excel)
                    cell_value = cell_value.replace('"', '""')
                    cell_value = f'"{cell_value}"'
                
                row_data.append(cell_value)
            copied_data.append("\t".join(row_data))
        
        clipboard_text = "\n".join(copied_data)
        clipboard = QApplication.clipboard()
        clipboard.setText(clipboard_text)
        
        self.info_label.setText(f"✓ Copiate {len(selected_rows)} righe dalla Lista {list_id} negli appunti")


    def show_duplicates_dialog(self, list_id):
        """Mostra una finestra con le righe duplicate"""
        store = self.lists[list_id]
        df = store["df"]
        
        # Filtra righe con chiave vuota PRIMA di cercare duplicati
        df_with_keys = self._get_df_with_valid_keys(df)
        
        if df_with_keys.empty:
            QMessageBox.information(self, "Nessun Dato", 
                                   f"La Lista {list_id} non contiene righe con chiave valida!")
            return
        
        # Trova le righe duplicate (ora senza le righe vuote)
        duplicated_mask = df_with_keys['KEY_NORMALIZED'].duplicated(keep=False)
        duplicates_df = df_with_keys[duplicated_mask].copy()
        
        if duplicates_df.empty:
            QMessageBox.information(self, "Nessun Duplicato", 
                                   f"La Lista {list_id} non contiene duplicati!")
            return
            
        # Crea finestra di dialogo
        dialog = QDialog(self)
        dialog.setWindowTitle(f"🔍 Duplicati - Lista {list_id}")
        dialog.resize(1000, 600)
        
        layout = QVBoxLayout(dialog)
        
        # Informazioni
        duplicates_unique = duplicates_df['KEY_NORMALIZED'].nunique()
        duplicates_count = len(duplicates_df)
        
        # Conta quante righe con chiave vuota sono state escluse
        empty_keys_count = len(df) - len(df_with_keys)
        
        info_text = (
            f"<b>Chiavi duplicate:</b> {duplicates_unique}<br>"
            f"<b>Righe totali duplicate:</b> {duplicates_count}"
        )
        
        if empty_keys_count > 0:
            info_text += f"<br><b>⚠️ Righe con chiave vuota (escluse):</b> {empty_keys_count}"
        
        info_label = QLabel(info_text)
        info_label.setStyleSheet("padding: 10px; background-color: #fff3cd; border-radius: 5px;")
        layout.addWidget(info_label)
        
        # Tabella con righe duplicate
        table = QTableWidget()
        table.setColumnCount(len(self.columns) + 1)  # +1 per la colonna "Riga"
        table.setHorizontalHeaderLabels(["Riga Visibile"] + self.columns)
        
        # Configura tabella
        table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        table.setAlternatingRowColors(True)
        table.setSortingEnabled(True)
        table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        
        # ✅ SOLUZIONE: Costruisci una mappa chiave -> posizione visibile nella tabella UI
        table_widget = store["table"]
        key_to_visual_row = {}
        
        for visual_row in range(table_widget.rowCount()):
            # Trova l'item della colonna chiave (con protezione da cambio preset)
            try:
                key_col_index = self.columns.index(self.key_column)
            except ValueError:
                QMessageBox.warning(
                    self, 
                    "⚠️ Errore Configurazione",
                    f"La colonna chiave '{self.key_column}' non esiste più.\n"
                    "Probabilmente hai cambiato preset dopo aver caricato i dati."
                )
                return
            item = table_widget.item(visual_row, key_col_index)
            
            if item:
                key_normalized = self.normalize_key(item.text())
                if key_normalized:
                    # Se la chiave ha duplicati, salva tutte le posizioni
                    if key_normalized not in key_to_visual_row:
                        key_to_visual_row[key_normalized] = []
                    key_to_visual_row[key_normalized].append(visual_row + 1)  # +1 per numerazione umana
        
        # Popola tabella (ordina per chiave normalizzata per raggruppare i duplicati)
        duplicates_sorted = duplicates_df.sort_values('KEY_NORMALIZED').reset_index(drop=False)
        table.setRowCount(len(duplicates_sorted))
        
        for i, row in duplicates_sorted.iterrows():
            key = row['KEY_NORMALIZED']
            
            # ✅ Usa la mappa per trovare le posizioni visibili
            visual_rows = key_to_visual_row.get(key, [])
            
            if visual_rows:
                # Se ci sono più righe con la stessa chiave, mostra tutte
                row_text = ", ".join(map(str, visual_rows))
                row_item = QTableWidgetItem(row_text)
                row_item.setBackground(QColor("#ffeaa7"))
                table.setItem(i, 0, row_item)
            else:
                # Fallback: la chiave non è più presente nella tabella UI
                # (può succedere se i dati sono stati modificati dopo il confronto)
                row_item = QTableWidgetItem("N/D")
                row_item.setBackground(QColor("#ffcccc"))
                row_item.setToolTip("Riga non trovata nella tabella corrente")
                table.setItem(i, 0, row_item)
            
            # Altre colonne
            for j, col_name in enumerate(self.columns):
                value = row[col_name] if col_name in row else ""
                item = QTableWidgetItem(str(value) if not pd.isna(value) else "")
                
                # Evidenzia la colonna chiave
                if col_name == self.key_column:
                    item.setBackground(QColor("#ff7675"))
                    font = QFont()
                    font.setBold(True)
                    item.setFont(font)
                
                table.setItem(i, j + 1, item)
        
        layout.addWidget(table)
        
        # Pulsante chiudi
        btn_close = QPushButton("Chiudi")
        btn_close.clicked.connect(dialog.accept)
        layout.addWidget(btn_close)
        
        dialog.exec()
        
        
    # ============================================================
    # MENU CONTESTUALE
    # ============================================================

    def show_context_menu(self, pos, list_id):
        """Menu contestuale UNIFICATO"""
        store = self.lists[list_id]
        menu = QMenu(self)
        
        # Azioni del menu
        paste_action = menu.addAction("📋 Incolla da Excel")
        copy_action = menu.addAction("📄 Copia righe selezionate")  # *** AGGIUNTO ***
        delete_action = menu.addAction("❌ Elimina righe selezionate")
        menu.addSeparator()
        restore_action = menu.addAction("↩️ Ripristina ultima eliminazione")
        
        # Disabilita "Copia" ed "Elimina" se non ci sono selezioni
        has_selection = bool(store["table"].selectedItems())
        copy_action.setEnabled(has_selection)  # *** AGGIUNTO ***
        delete_action.setEnabled(has_selection)
        
        # Disabilita "Ripristina" se non ci sono eliminazioni da ripristinare
        restore_action.setEnabled(len(store["deleted"]) > 0)
        
        # Connetti le azioni (CENTRALIZZATO)
        paste_action.triggered.connect(lambda: self.paste_data(list_id))
        copy_action.triggered.connect(lambda: self.copy_selected_rows(list_id))  # *** AGGIUNTO ***
        delete_action.triggered.connect(lambda: self.delete_rows(list_id))
        restore_action.triggered.connect(lambda: self.restore_rows(list_id))
        
        # Mostra il menu nella posizione del cursore
        menu.exec(store["table"].viewport().mapToGlobal(pos))


    def _get_df_with_valid_keys(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Filtra il DataFrame escludendo righe con chiavi vuote o None.
        
        Args:
            df: DataFrame con colonna KEY_NORMALIZED
        
        Returns:
            DataFrame filtrato con solo chiavi valide
        
        Example:
            >>> df = pd.DataFrame({'KEY_NORMALIZED': ['ABC123', None, '', 'XYZ789']})
            >>> valid_df = self._get_df_with_valid_keys(df)
            >>> len(valid_df)  # Restituisce 2 (solo ABC123 e XYZ789)
        """
        if df.empty or 'KEY_NORMALIZED' not in df.columns:
            return pd.DataFrame()
        
        return df[
            df['KEY_NORMALIZED'].notna() & 
            (df['KEY_NORMALIZED'].astype(str).str.strip() != "")
        ].copy()


    # ============================================================
    # UTILITÀ
    # ============================================================

    def normalize_key(self, key_value) -> str:
        """
        Normalizza la colonna chiave rimuovendo caratteri non alfanumerici.
        
        Args:
            key_value: Valore da normalizzare
        
        Returns:
            Stringa normalizzata (solo caratteri alfanumerici in maiuscolo)
            Restituisce stringa vuota se il valore è None/NaN o vuoto dopo la normalizzazione
        """
        if pd.isna(key_value):
            return ""
        
        key_str = str(key_value).strip()
        
        # ✅ MIGLIORAMENTO: Gestisci esplicitamente stringhe vuote
        if not key_str:
            return ""
        
        normalized = ''.join(c for c in key_str if c.isalnum()).upper()
        
        # ✅ MIGLIORAMENTO: Log per debug se la normalizzazione rimuove tutto
        if not normalized and key_str:
            logging.debug(f"Chiave '{key_str}' normalizzata in stringa vuota")
        
        return normalized

    
    def _get_display_columns(self) -> list:
        """Restituisce solo le colonne visibili (esclude colonne tecniche)"""
        return [col for col in self.columns if not col.startswith('__')]
        
    def _parse_clipboard_text(self, text):
        """
        Trasforma il testo della clipboard in DataFrame grezzo.
        ✅ LEGGE TUTTE LE COLONNE senza troncare in base al preset.
        """
        try:
            # Usa csv.reader per gestire virgolette
            rows = []
            reader = csv.reader(
                io.StringIO(text), 
                delimiter='\t',
                quotechar='"',
                doublequote=True,
                skipinitialspace=False
            )
            
            for row in reader:
                rows.append(row)
            
            if not rows:
                raise pd.errors.EmptyDataError("Nessuna riga trovata")
            
            # ✅ TROVA IL NUMERO MASSIMO DI COLONNE NEI DATI INCOLLATI
            max_cols = max(len(row) for row in rows)
            
            # ✅ NORMALIZZA TUTTE LE RIGHE ALLA STESSA LUNGHEZZA
            processed_rows = []
            for row in rows:
                # Estendi righe corte con stringhe vuote
                normalized_row = row + [''] * (max_cols - len(row))
                processed_rows.append(normalized_row)
            
            if not processed_rows:
                raise pd.errors.EmptyDataError("Nessuna riga valida trovata")
            
            # ✅ CREA DATAFRAME CON TUTTE LE COLONNE
            df = pd.DataFrame(processed_rows, dtype=str)
            
            # ✅ Traccia il numero di riga originale
            df['__ORIGINAL_ROW__'] = range(1, len(df) + 1)
            
            # ✅ GENERA NOMI COLONNE TEMPORANEI O USA PRESET
            if max_cols == len(self.columns):
                # Stesso numero di colonne: usa preset
                column_names = self.columns + ['__ORIGINAL_ROW__']
            elif max_cols < len(self.columns):
                # Meno colonne: usa solo le prime del preset
                column_names = self.columns[:max_cols] + ['__ORIGINAL_ROW__']
            else:
                # Più colonne: usa preset + colonne extra generiche
                extra_cols = [f"Colonna_{i+1}" for i in range(len(self.columns), max_cols)]
                column_names = self.columns + extra_cols + ['__ORIGINAL_ROW__']
            
            df.columns = column_names
            
            return df
            
        except pd.errors.EmptyDataError:
            QMessageBox.warning(self, "Dati Vuoti", 
                               "I dati copiati sono vuoti o non validi.")
            return None
            
        except Exception as e:
            QMessageBox.critical(
                self, 
                "❌ Errore Critico - Parsing",
                f"Errore nel parsing dei dati:\n\n{str(e)}"
            )
            logging.exception(f"Errore parsing clipboard: {e}")
            return None
        
    
    def _handle_header_detection(self, df: pd.DataFrame) -> pd.DataFrame | None:
        """
        Gestisce il rilevamento automatico dell'intestazione.
        
        Args:
            df: DataFrame da validare
        
        Returns:
            DataFrame con intestazione applicata, o None se l'utente annulla
        
        Note:
            Questo metodo:
            1. Controlla se la prima riga sembra un'intestazione
            2. Chiede conferma all'utente
            3. Applica l'intestazione se confermato
        """
        if df.empty:
            return df
        
        first_row_key = df.iloc[0][self.key_column]
        first_row_key_norm = self.normalize_key(first_row_key)
        
        # Determina se la prima riga è un'intestazione
        is_header = (
            not first_row_key_norm or
            first_row_key_norm.isalpha() or
            first_row_key_norm == self.normalize_key(self.key_column)
        )
        
        if not is_header:
            return df  # Non è un'intestazione, continua
        
        # Chiedi conferma all'utente
        reply = QMessageBox.question(
            self,
            "📋 Intestazione Rilevata",
            f"La prima riga sembra essere un'intestazione:\n"
            f"Valore colonna chiave: '{first_row_key}'\n\n"
            f"Vuoi usare la prima riga come intestazione ed eliminarla dai dati?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.No:
            return None  # Utente ha annullato
        
        # Applica intestazione
        return self._apply_header_from_first_row(df)
        
        
    def _apply_header_from_first_row(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Applica l'intestazione dalla prima riga del DataFrame.
        ✅ Usa TUTTE le colonne del DataFrame, non solo quelle del preset.
        """
        # ✅ Estrai nomi da TUTTE le colonne del DataFrame (esclusa __ORIGINAL_ROW__)
        data_columns = [col for col in df.columns if col != '__ORIGINAL_ROW__']
        
        new_columns = []
        for i, col_name in enumerate(data_columns):
            # Prendi il valore dalla prima riga
            value = df.iloc[0][col_name]
            if not pd.isna(value) and str(value).strip():
                new_columns.append(str(value).strip())
            else:
                # Fallback: mantieni il nome corrente
                new_columns.append(col_name)
        
        # Rimuovi prima riga e aggiorna colonne
        df = df.iloc[1:].reset_index(drop=True)
        df.columns = new_columns + ['__ORIGINAL_ROW__']
        
        # ✅ Aggiorna colonna chiave (cerca il nome nell'intestazione)
        try:
            old_key_idx = list(self.presets[self.current_preset]["columns"]).index(self.key_column)
            if old_key_idx < len(new_columns):
                self.key_column = new_columns[old_key_idx]
            else:
                # Fallback: usa la prima colonna
                self.key_column = new_columns[0] if new_columns else self.key_column
        except (ValueError, IndexError):
            self.key_column = new_columns[0] if new_columns else self.key_column
        
        # ✅ AGGIORNA LE COLONNE GLOBALI
        self.columns = new_columns
        self.table_manager.columns = new_columns
        
        # Aggiorna tutte le tabelle
        for list_id in [1, 2]:
            table = self.lists[list_id]["table"]
            table.setColumnCount(len(self.columns))
            table.setHorizontalHeaderLabels(self.columns)
            self.table_manager.enable_manual_sorting(table)  # ✅ RIABILITA SORTING
        
        for result_table in [self.res_matches, self.res_mismatches]:
            result_table.setColumnCount(len(self.columns))
            result_table.setHorizontalHeaderLabels(self.columns)
            self.table_manager.enable_manual_sorting(result_table)  # ✅ RIABILITA SORTING
            self.table_manager.enable_manual_sorting(result_table)  # ✅ Riabilita sorting
        
        # ✅ CANCELLA I RISULTATI OBSOLETI (le colonne sono cambiate!)
        if self.matches_df is not None or self.mismatches_df is not None:
            self.clear_results(confirm=False)
        
        self.info_label.setText(
            f"✅ Intestazione applicata: colonna chiave = '{self.key_column}'"
        )
        
        return df
        
    def _validate_imported_df(self, df: pd.DataFrame) -> pd.DataFrame | None:
        """
        Valida il DataFrame importato e normalizza la colonna chiave.
        Gestisce automaticamente la presenza di intestazioni nella prima riga.
        
        Args:
            df: DataFrame da validare
        
        Returns:
            DataFrame validato con KEY_NORMALIZED, o None se validazione fallisce
        """
        # Verifica che la colonna chiave esista
        if self.key_column not in df.columns:
            QMessageBox.warning(
                self,
                "Colonna Chiave Mancante",
                f"I dati incollati non contengono la colonna chiave '{self.key_column}'.\n"
                f"Verifica di aver selezionato il preset corretto nelle Impostazioni."
            )
            return None
        
        # Gestione intestazione
        df = self._handle_header_detection(df)
        if df is None:
            return None  # Utente ha annullato
        
        # Normalizza colonna chiave per tutte le righe
        df['KEY_NORMALIZED'] = df[self.key_column].apply(self.normalize_key)
        
        # Verifica chiavi vuote
        empty_mask = df['KEY_NORMALIZED'].isna() | (df['KEY_NORMALIZED'] == "")
        if empty_mask.any():
            empty_count = empty_mask.sum()
            reply = QMessageBox.question(
                self,
                "⚠️ Chiavi Vuote Rilevate",
                f"Attenzione: {empty_count} righe hanno la colonna chiave '{self.key_column}' vuota.\n\n"
                f"Queste righe potrebbero causare problemi nel confronto.\n"
                f"Vuoi continuare comunque?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            if reply == QMessageBox.StandardButton.No:
                return None
        
        return df
    
    
    
    
    def get_clipboard_data(self) -> pd.DataFrame | None:
        """Legge la clipboard e la trasforma in DataFrame validato"""
        clipboard = QApplication.clipboard()
        text = clipboard.text()
        
        if not text or not text.strip():
            QMessageBox.warning(self, "Clipboard Vuota", 
                              "La clipboard è vuota. Copia prima i dati da Excel.")
            return None
        
        # Parsing separato dalla validazione
        df = self._parse_clipboard_text(text)
        if df is None:
            return None
        
        # Validazione separata dal parsing
        df = self._validate_imported_df(df)
        
        return df

    def update_table_display(self, table_widget: QTableWidget, df: pd.DataFrame, 
                            highlight_from: int | None = None) -> None:
        """Popola la QTableWidget con i dati del DataFrame"""
        display_columns = [col for col in self.columns if col in df.columns and not col.startswith('__')]
        display_df = df[display_columns]
        
        show_cursor = len(display_df) > 1000
        if show_cursor:
            QApplication.setOverrideCursor(Qt.CursorShape.WaitCursor)
        
        # Salva stato sorting e disabilitalo temporaneamente
        was_sorting_enabled = table_widget.isSortingEnabled()
        table_widget.setSortingEnabled(False)
        table_widget.setUpdatesEnabled(False)
        
        try:
            table_widget.setRowCount(len(display_df))
            
            is_res = table_widget in [self.res_matches, self.res_mismatches]
            bg_color = QColor(MATCH_COLOR) if table_widget == self.res_matches else QColor(MISMATCH_COLOR)
            color_highlight = QColor(HIGHLIGHT_COLOR)
            color_white = QColor(WHITE_COLOR)
            
            # ✅ OTTIMIZZAZIONE: Pre-alloca tutti gli item in una volta sola
            for i in range(len(display_df)):
                for j in range(len(display_columns)):
                    if not table_widget.item(i, j):
                        table_widget.setItem(i, j, QTableWidgetItem())

            # ✅ OTTIMIZZAZIONE: Popola i dati senza ricreare gli item
            for i, row in enumerate(display_df.itertuples(index=True)):
                real_index = row.Index  # ✅ Usa l'indice esplicito del DataFrame
                for j, col_name in enumerate(display_columns):
                    # ✅ Accesso per indice (già protetto contro IndexError)
                    value = row[j + 1] if j + 1 < len(row) else ""
                    
                    item = table_widget.item(i, j)
                    val_str = str(value) if not pd.isna(value) else ""
                    item.setText(val_str)
                    item.setToolTip(val_str if len(val_str) > 30 else "")
                    item.setData(Qt.ItemDataRole.UserRole, real_index)
                    
                    if highlight_from is not None and i >= highlight_from:
                        item.setBackground(color_highlight)
                    elif is_res:
                        item.setBackground(bg_color)
                    else:
                        item.setBackground(color_white)
        
        except Exception as e:
            QMessageBox.critical(
                self,
                "⌠Errore Critico - Visualizzazione",
                f"Impossibile visualizzare i dati:\n\n{str(e)}"
            )
            logging.exception(f"Errore in update_table_display: {e}")
        
        finally:
            table_widget.setUpdatesEnabled(True)
            # Ripristina stato sorting originale
            table_widget.setSortingEnabled(was_sorting_enabled)
            
            if show_cursor:
                QApplication.restoreOverrideCursor()

    # ============================================================
    # CONFRONTO E RISULTATI
    # ============================================================

    def compare_data(self) -> None:
        """Confronta le due liste"""
        if self.lists[1]["df"].empty or self.lists[2]["df"].empty:
            QMessageBox.warning(self, "Dati Mancanti", 
                              "Devi caricare entrambe le liste prima di confrontarle!")
            return

        QApplication.setOverrideCursor(Qt.CursorShape.WaitCursor)
        self.info_label.setText("⏳ Confronto in corso...")
        QApplication.processEvents()
        
        logging.info(f"Inizio confronto: Lista1={len(self.lists[1]['df'])} righe, Lista2={len(self.lists[2]['df'])} righe")
        
        try:
            matches, mismatches, stats = ComparisonEngine.compare(
                self.lists[1]["df"],
                self.lists[2]["df"]
            )
            
            # Verifica duplicati con conferma utente
            if stats['duplicates_df1'] > 0:
                reply = QMessageBox.question(
                    self, 
                    "⚠️ Duplicati Rilevati",
                    f"Attenzione: ci sono {stats['duplicates_df1']} chiavi duplicate nella Lista 1.\n"
                    f"Questo potrebbe alterare il conteggio dei mancanti.\n\n"
                    f"Vuoi continuare comunque?",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
                )
                if reply == QMessageBox.StandardButton.No:
                    return
            
            if stats['duplicates_df2'] > 0:
                reply = QMessageBox.question(
                    self,
                    "⚠️ Duplicati Rilevati",
                    f"Attenzione: ci sono {stats['duplicates_df2']} chiavi duplicate nella Lista 2.\n"
                    f"Questo potrebbe causare conteggi errati nelle corrispondenze.\n\n"
                    f"Vuoi continuare comunque?",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
                )
                if reply == QMessageBox.StandardButton.No:
                    return

            self.matches_df = matches
            self.mismatches_df = mismatches

            self.update_table_display(self.res_matches, matches)
            self.update_table_display(self.res_mismatches, mismatches)
            
            self.tabs_results.setTabText(0, f"✓ Corrispondenze ({len(matches)})")
            self.tabs_results.setTabText(1, f"✗ Mancanti ({len(mismatches)})")
            
            self.btn_export_matches.setEnabled(len(matches) > 0)
            self.btn_export_mismatches.setEnabled(len(mismatches) > 0)
            self.btn_clear_results.setEnabled(True)
            
            self.info_label.setText(
                f"✓ Confronto completato: {len(matches)} corrispondenze, "
                f"{len(mismatches)} mancanti su {len(self.lists[1]['df'])} totali"
            )
            
            QMessageBox.information(self, "Confronto Completato", 
                f"Risultati:\n\n"
                f"✓ Corrispondenze: {len(matches)}\n"
                f"✗ Mancanti: {len(mismatches)}\n"
                f"📊 Totale Lista 1: {len(self.lists[1]['df'])}"
            )
            
        except KeyError as e:
            QMessageBox.critical(
                self, 
                "⌠Errore Critico - Configurazione", 
                f"Colonna mancante nel confronto:\n\n{str(e)}\n\n"
                f"Verifica che la configurazione corrisponda ai dati caricati."
            )
            logging.exception(f"Errore chiave mancante: {e}")
        except Exception as e:
            QMessageBox.critical(
                self, 
                "⌠Errore Critico - Confronto", 
                f"Errore durante il confronto:\n\n{str(e)}"
            )
            logging.exception(f"Errore in compare_data: {e}")
        finally:
            QApplication.restoreOverrideCursor()

    def clear_results(self, confirm: bool = True) -> None:
        """Cancella i risultati del confronto
        
        Args:
            confirm: Se True, chiede conferma all'utente (default: True)
        """
        if confirm:
            reply = QMessageBox.question(
                self, 
                "Conferma Cancellazione",
                "Vuoi cancellare i risultati del confronto?\n\n"
                "Le liste originali (Lista 1 e Lista 2) non verranno modificate.",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No
            )
            
            if reply != QMessageBox.StandardButton.Yes:
                return
        
        self.res_matches.setRowCount(0)
        self.res_mismatches.setRowCount(0)
        
        self.tabs_results.setTabText(0, "✓ Corrispondenze (0)")
        self.tabs_results.setTabText(1, "✗ Mancanti (0)")
        
        self.btn_export_matches.setEnabled(False)
        self.btn_export_mismatches.setEnabled(False)
        self.btn_clear_results.setEnabled(False)
        
        self.matches_df = None
        self.mismatches_df = None
        
        self.info_label.setText("✓ Risultati cancellati. Pronto per un nuovo confronto.")
            
            
    def swap_lists(self) -> None:
        """Inverte Lista 1 con Lista 2"""
        # Verifica se almeno una lista ha dati
        if self.lists[1]["df"].empty and self.lists[2]["df"].empty:
            QMessageBox.information(self, "Nessun Dato", 
                                  "Entrambe le liste sono vuote. Nulla da invertire!")
            return
        
        # Chiedi conferma
        reply = QMessageBox.question(
            self,
            "🔄 Conferma Inversione",
            "Vuoi invertire Lista 1 con Lista 2?\n\n"
            "• Lista 1 (Worklist) diventerà Lista 2 (Confronto)\n"
            "• Lista 2 (Confronto) diventerà Lista 1 (Worklist)\n\n"
            "I risultati del confronto verranno cancellati.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.No:
            return
        
        # Scambia i DataFrame
        temp_df = self.lists[1]["df"].copy()
        self.lists[1]["df"] = self.lists[2]["df"].copy()
        self.lists[2]["df"] = temp_df
        
        # Scambia lo storico eliminazioni
        temp_deleted = self.lists[1]["deleted"].copy()
        self.lists[1]["deleted"] = self.lists[2]["deleted"].copy()
        self.lists[2]["deleted"] = temp_deleted
        
        # Aggiorna le tabelle
        for list_id in [1, 2]:
            store = self.lists[list_id]
            self.update_table_display(store["table"], store["df"])
            self._update_counters(list_id)
            store["btn_restore"].setEnabled(len(store["deleted"]) > 0)
        
        # Cancella i risultati
        self.clear_results(confirm=False)
        
        self.info_label.setText("✓ Liste invertite con successo")
        
        QMessageBox.information(self, "✓ Inversione Completata",
            "Le liste sono state invertite:\n\n"
            f"• Lista 1: {len(self.lists[1]['df'])} righe\n"
            f"• Lista 2: {len(self.lists[2]['df'])} righe")
    

    # ============================================================
    # ESPORTAZIONE UNIFICATA
    # ============================================================

    def export_results(self, result_type: str) -> None:
        """
        Metodo UNIFICATO per esportare corrispondenze o mancanti.
        
        Args:
            result_type: "matches" o "mismatches"
        """
        df = self.matches_df if result_type == "matches" else self.mismatches_df
        prefix = f"{self.key_column}_Corrispondenze" if result_type == "matches" else f"{self.key_column}_Mancanti"
        sheet_name = "Corrispondenze" if result_type == "matches" else "Mancanti"
        
        if df is None or df.empty:
            QMessageBox.warning(self, "Nessun Dato", 
                              f"Non ci sono {sheet_name.lower()} da esportare!")
            return
        
        desktop = QStandardPaths.writableLocation(QStandardPaths.StandardLocation.DesktopLocation)
        file_path = self._get_unique_filepath(desktop, prefix, "xlsx")
        
        try:
            self._export_to_formatted_excel(
                df[self.columns], 
                file_path,
                sheet_name
            )
            
            QMessageBox.information(self, "Esportazione Completata", 
                f"File salvato con successo sul Desktop:\n{os.path.basename(file_path)}")
            self.info_label.setText(f"✓ Esportati {len(df)} {sheet_name.lower()}: {os.path.basename(file_path)}")
            
        except Exception as e:
            QMessageBox.critical(
                self, 
                "⌠Errore Critico - Esportazione",
                f"Errore durante l'esportazione:\n\n{str(e)}"
            )
            logging.exception(f"Errore in export_results: {e}")

    def _get_unique_filepath(self, directory, base_name, extension):
        """
        Genera un percorso file univoco con numerazione progressiva.
        
        Args:
            directory: Cartella di destinazione
            base_name: Nome base del file (senza estensione)
            extension: Estensione del file (senza punto)
        
        Returns:
            Percorso completo del file con numerazione progressiva se necessario
        """
        # Prova prima senza numero
        filepath = os.path.join(directory, f"{base_name}.{extension}")
        
        if not os.path.exists(filepath):
            return filepath
        
        # Se esiste, aggiungi numerazione progressiva
        counter = 1
        while True:
            filepath = os.path.join(directory, f"{base_name}_{counter}.{extension}")
            if not os.path.exists(filepath):
                return filepath
            counter += 1

    @staticmethod
    def _calculate_excel_column_width(column_name):
        """
        Calcola la larghezza ottimale per una colonna Excel basandosi sul nome.
        La ricerca è case-insensitive.
        
        Args:
            column_name: Nome della colonna (es. "ISBN", "Titolo", "autore")
        
        Returns:
            int: Larghezza in unità Excel (default 20)
        """
        # ✅ Gestisce None e valori non-stringa
        col_name_upper = str(column_name).upper().strip()
        
        for target_width, keywords in EXCEL_COLUMN_WIDTHS.items():
            # ✅ Ricerca bidirezionale: keyword nella colonna O colonna nella keyword
            if any(kw in col_name_upper or col_name_upper in kw for kw in keywords):
                return target_width
        
        return 20  # Default


    def _export_to_formatted_excel(self, df: pd.DataFrame, filepath: str, sheet_name: str):
        """
        Esporta DataFrame in Excel con formattazione professionale.
        
        Specifiche di formattazione:
        - Header azzurro (#8DBEE3) con testo in grassetto
        - Larghezze colonne ottimizzate dinamicamente
        - Altezza righe uniforme (30px)
        - Allineamento verticale top, orizzontale left
        - Wrap text abilitato
        - Layout stampa ottimizzato (A4 orizzontale)
        - Freeze panes sulla prima riga
        - Zoom 110%
        
        Args:
            df: DataFrame da esportare
            filepath: Percorso file di output
            sheet_name: Nome del foglio Excel
        """
        # Crea workbook e worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # ====================================================================
        # CONFIGURAZIONE LAYOUT E STAMPA
        # ====================================================================
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.freeze_panes = 'A2'
        ws.sheet_view.zoomScale = DEFAULT_ZOOM_LEVEL

        
        # ====================================================================
        # STILI
        # ====================================================================
        header_fill = PatternFill(start_color='8DBEE3', end_color='8DBEE3', fill_type='solid')
        header_font = Font(bold=True)
        alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
        
        # ====================================================================
        # SCRITTURA HEADER CON LARGHEZZE DINAMICHE
        # ====================================================================
        for col_idx, column_name in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=column_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = alignment
            
            # Calcolo larghezza intelligente basato sul nome colonna
            width = self._calculate_excel_column_width(column_name)
            ws.column_dimensions[get_column_letter(col_idx)].width = width
            
        
        # Altezza riga header
        ws.row_dimensions[1].height = 19
        
    
        # ====================================================================
        # SCRITTURA DATI CON ALTEZZA UNIFORME (CICLO UNIFICATO)
        # ====================================================================
        # Calcola il numero totale di righe da formattare (minimo 100 per consistenza visiva)
        total_rows = max(len(df) + 1, DEFAULT_EXCEL_MIN_ROWS)


        # Imposta l'altezza per TUTTE le righe (anche quelle vuote)
        for row_idx in range(2, total_rows + 1):
            ws.row_dimensions[row_idx].height = 30

        # Scrivi i dati nelle righe valorizzate
        for row_idx, row_data in enumerate(df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value if not pd.isna(value) else ""
                cell.alignment = alignment
        
        # ====================================================================
        # SALVATAGGIO
        # ====================================================================
        wb.save(filepath)

    # ============================================================
    # IMPOSTAZIONI E PRESET
    # ============================================================

    def open_settings(self):
        """Apre la finestra di dialogo delle impostazioni"""
        # Controlla se ci sono dati caricati
        has_data = not self.lists[1]["df"].empty or not self.lists[2]["df"].empty
        has_results = self.matches_df is not None or self.mismatches_df is not None
        
        if has_data or has_results:
            reply = QMessageBox.warning(
                self,
                "⚠️ Attenzione",
                "Cambiare configurazione cancellerà TUTTI i dati caricati e i risultati.\n\n"
                "Vuoi continuare?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No
            )
            
            if reply != QMessageBox.StandardButton.Yes:
                return
        
        dialog = SettingsDialog(self.presets, self.current_preset, self.default_preset, self.protected_presets, self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            result = dialog.get_result()
            
            if result['action'] == 'select':
                # Carica preset esistente
                self.current_preset = result['preset_name']
                self.key_column = self.presets[self.current_preset]["key_column"]
                self.columns = self.presets[self.current_preset]["columns"].copy()
                self.table_manager.columns = self.columns  # ✅ SINCRONIZZA
                
            elif result['action'] == 'create':
                # Crea nuovo preset
                preset_name = result['preset_name']
                self.presets[preset_name] = {
                    "key_column": result['key_column'],
                    "columns": result['columns']
                }
                self.current_preset = preset_name
                self.key_column = result['key_column']
                self.columns = result['columns']
                self.table_manager.columns = self.columns  # ✅ SINCRONIZZA
                
                # *** NUOVO: Imposta come predefinito se richiesto ***
                if result.get('set_as_default', False):
                    self.default_preset = preset_name
                
                # --- SALVA AUTOMATICAMENTE SU DISCO ---
                self.save_presets()
            
            # *** NUOVO: Gestisci azione "set_default" ***
            elif result['action'] == 'set_default':
                self.default_preset = result['preset_name']
                self.save_presets()
                QMessageBox.information(
                    self,
                    "✓ Preset Predefinito Impostato",
                    f"Il preset '{result['preset_name']}' è ora il predefinito.\n\n"
                    "Verrà caricato automaticamente all'avvio dell'applicazione."
                )
                return  # *** Non resettare i dati, solo salva il predefinito ***
            
                       
                        
            # ✅ SINCRONIZZA I PRESET (potrebbero essere stati eliminati nel dialog)
            self.presets = dialog.presets.copy()
            self.default_preset = dialog.default_preset
            
            # Resetta le liste
            self.reset_all_data()
            
            # Aggiorna il titolo della finestra
            self.setWindowTitle(f"Code Comparator - {self.current_preset}")
            
            self.info_label.setText(f"✓ Configurazione cambiata: {self.current_preset} (Colonna chiave: {self.key_column})")


    def _reset_table_widget(self, table_widget):
        """Resetta una tabella (funzione di aiuto per non ripetere codice)"""
        table_widget.setColumnCount(len(self.columns))
        table_widget.setHorizontalHeaderLabels(self.columns)
        self.table_manager.reset_sorting(table_widget)
        table_widget.setRowCount(0)
    
    def _reset_single_list(self, list_id: int) -> None:
        """Reset di una singola lista (helper method)"""
        store = self.lists[list_id]
        store["df"] = pd.DataFrame(columns=self.columns)
        store["deleted"].clear()
        
        # Reset tabella - USA LA NUOVA FUNZIONE!
        self._reset_table_widget(store["table"])  # ← UNA RIGA INVECE DI 4!
        
        # Reset contatori
        store["label"].setText("(0 righe)")
        store["duplicates_label"].setText("")
        
        # Reset pulsante ripristina
        store["btn_restore"].setEnabled(False)

    def reset_all_data(self):
        """Reset completo di tutti i dati quando si cambia configurazione"""
        # Reset liste usando helper method
        for list_id in [1, 2]:
            self._reset_single_list(list_id)
        
        # Reset risultati - USA LA NUOVA FUNZIONE!
        for table in [self.res_matches, self.res_mismatches]:
            self._reset_table_widget(table)  # ← UNA RIGA INVECE DI 2!
        
        # Reset tab
        self.tabs_results.setTabText(0, "✅ Corrispondenze (0)")
        self.tabs_results.setTabText(1, "✗ Mancanti (0)")
            
        # Disabilita pulsanti export
        self.btn_export_matches.setEnabled(False)
        self.btn_export_mismatches.setEnabled(False)
        self.btn_clear_results.setEnabled(False)
        
        # Rimuovi dataframe risultati
        self.matches_df = None
        self.mismatches_df = None

    

    def apply_styles(self):
        """Applica gli stili CSS all'applicazione"""
        self.setStyleSheet(APP_STYLESHEET)
        
        # Imposta gli object names per i pulsanti
        self.btn_swap.setObjectName("btn_swap")
        self.btn_compare.setObjectName("btn_compare")
        self.btn_export_matches.setObjectName("btn_export_matches")
        self.btn_export_mismatches.setObjectName("btn_export_mismatches")
        self.btn_clear_results.setObjectName("btn_clear_results")

# ============================================================
# MAIN
# ============================================================

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = CodeComparator()
    window.show()
    sys.exit(app.exec())
