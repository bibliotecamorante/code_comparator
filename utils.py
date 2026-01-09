"""
Funzioni di utilità generiche
"""
import os
import logging
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter

from config import EXCEL_COLUMN_WIDTHS, DEFAULT_EXCEL_MIN_ROWS, DEFAULT_ZOOM_LEVEL


def normalize_key(key_value) -> str:
    """
    Normalizza la colonna chiave rimuovendo caratteri non alfanumerici.
    
    Args:
        key_value: Valore da normalizzare
    
    Returns:
        Stringa normalizzata (solo caratteri alfanumerici in maiuscolo)
        Restituisce stringa vuota se il valore è None/NaN o vuoto dopo la normalizzazione
    """
    if pd.isna(key_value):
        return ""
    
    key_str = str(key_value).strip()
    
    if not key_str:
        return ""
    
    normalized = ''.join(c for c in key_str if c.isalnum()).upper()
    
    if not normalized and key_str:
        logging.debug(f"Chiave '{key_str}' normalizzata in stringa vuota")
    
    return normalized


def calculate_excel_column_width(column_name):
    """
    Calcola la larghezza ottimale per una colonna Excel basandosi sul nome.
    La ricerca è case-insensitive.
    
    Args:
        column_name: Nome della colonna (es. "ISBN", "Titolo", "autore")
    
    Returns:
        int: Larghezza in unità Excel (default 20)
    """
    col_name_upper = str(column_name).upper().strip()
    
    for target_width, keywords in EXCEL_COLUMN_WIDTHS.items():
        if any(kw in col_name_upper or col_name_upper in kw for kw in keywords):
            return target_width
    
    return 20


def get_unique_filepath(directory, base_name, extension):
    """
    Genera un percorso file univoco con numerazione progressiva.
    
    Args:
        directory: Cartella di destinazione
        base_name: Nome base del file (senza estensione)
        extension: Estensione del file (senza punto)
    
    Returns:
        Percorso completo del file con numerazione progressiva se necessario
    """
    filepath = os.path.join(directory, f"{base_name}.{extension}")
    
    if not os.path.exists(filepath):
        return filepath
    
    counter = 1
    while True:
        filepath = os.path.join(directory, f"{base_name}_{counter}.{extension}")
        if not os.path.exists(filepath):
            return filepath
        counter += 1


def export_to_formatted_excel(df: pd.DataFrame, filepath: str, sheet_name: str):
    """
    Esporta DataFrame in Excel con formattazione professionale.
    
    Specifiche di formattazione:
    - Header azzurro (#8DBEE3) con testo in grassetto
    - Larghezze colonne ottimizzate dinamicamente
    - Altezza righe uniforme (30px)
    - Allineamento verticale top, orizzontale left
    - Wrap text abilitato
    - Layout stampa ottimizzato (A4 orizzontale)
    - Freeze panes sulla prima riga
    - Zoom 110%
    
    Args:
        df: DataFrame da esportare
        filepath: Percorso file di output
        sheet_name: Nome del foglio Excel
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # Configurazione layout e stampa
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.freeze_panes = 'A2'
    ws.sheet_view.zoomScale = DEFAULT_ZOOM_LEVEL

    # Stili
    header_fill = PatternFill(start_color='8DBEE3', end_color='8DBEE3', fill_type='solid')
    header_font = Font(bold=True)
    alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
    
    # Scrittura header con larghezze dinamiche
    for col_idx, column_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=column_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = alignment
        
        width = calculate_excel_column_width(column_name)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    ws.row_dimensions[1].height = 19

    # Scrittura dati con altezza uniforme
    total_rows = max(len(df) + 1, DEFAULT_EXCEL_MIN_ROWS)

    for row_idx in range(2, total_rows + 1):
        ws.row_dimensions[row_idx].height = 30

    for row_idx, row_data in enumerate(df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value if not pd.isna(value) else ""
            cell.alignment = alignment
    
    wb.save(filepath)